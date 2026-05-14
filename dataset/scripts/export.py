#!/usr/bin/env python3
"""
export.py — Genera los JSON de src/data/ a partir de dataset/datosfinales.xlsx.

Uso:
    python3 dataset/scripts/export.py
    python3 dataset/scripts/export.py --dry-run
"""

import re, sys, json, os
from datetime import datetime, timedelta, date
from typing import Optional
import openpyxl

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
DATASET_DIR = os.path.dirname(SCRIPT_DIR)
ROOT_DIR    = os.path.dirname(DATASET_DIR)
XLSX        = os.path.join(DATASET_DIR, "datosfinales.xlsx")
DATA_DIR    = os.path.join(ROOT_DIR, "src", "data")

C_CAT=1; C_NOMBRE=2; C_TW=3; C_TW_A=4; C_BS=5; C_BS_A=6
C_MD=7; C_MD_A=8; C_EMAIL=9; C_DETALLE=10; C_TIPO=11
C_GRUPO=12; C_CIRCUN=13; C_CCAA=14; C_PARTIDO=15

def fmt_date(val):
    if val is None: return None
    if isinstance(val, datetime): return val.strftime("%Y-%m-%d")
    s = str(val).strip()
    if not s: return None
    if s == "404": return "404"
    if re.match(r"^\d{4}-\d{2}-\d{2}", s): return s[:10]
    return None

def load_refs():
    path = os.path.join(DATA_DIR, "lastUpdate.json")
    try:
        with open(path, encoding="utf-8") as f: return json.load(f)
    except: return {}

def is_active(val, ref_date=None):
    if not isinstance(val, str): return False
    m = re.match(r"^(\d{4}-\d{2}-\d{2})", val)
    if not m: return False
    try:
        d = datetime.strptime(m.group(1), "%Y-%m-%d").date()
        ref = ref_date or datetime.now().date()
        return (ref - timedelta(days=30)) <= d <= ref
    except ValueError: return False

def v(ws, row, col):
    val = ws.cell(row, col).value
    return None if val is None or val == "" else val

def s(ws, row, col):
    val = v(ws, row, col)
    if val is None: return None
    r = str(val).strip()
    return r if r else None

def build_age(ws, rows):
    return [{"nombre": s(ws,r,C_NOMBRE), "categoria": s(ws,r,C_DETALLE),
             "twitter": s(ws,r,C_TW), "twitter_activo": fmt_date(v(ws,r,C_TW_A)),
             "bluesky": s(ws,r,C_BS), "bluesky_activo": fmt_date(v(ws,r,C_BS_A)),
             "mastodon": s(ws,r,C_MD), "mastodon_activo": fmt_date(v(ws,r,C_MD_A)),
             "email": s(ws,r,C_EMAIL)} for r in rows]

def build_autonomias(ws, rows):
    return [{"tipo": s(ws,r,C_TIPO), "ccaa": s(ws,r,C_CCAA), "nombre": s(ws,r,C_NOMBRE),
             "partido": s(ws,r,C_PARTIDO),
             "twitter": s(ws,r,C_TW), "twitter_activo": fmt_date(v(ws,r,C_TW_A)),
             "bluesky": s(ws,r,C_BS), "bluesky_activo": fmt_date(v(ws,r,C_BS_A)),
             "mastodon": s(ws,r,C_MD), "mastodon_activo": fmt_date(v(ws,r,C_MD_A)),
             "email": s(ws,r,C_EMAIL)} for r in rows]

def build_gobierno(ws, rows):
    return [{"nombre": s(ws,r,C_NOMBRE), "cargo": s(ws,r,C_DETALLE),
             "twitter": s(ws,r,C_TW), "twitter_activo": fmt_date(v(ws,r,C_TW_A)),
             "bluesky": s(ws,r,C_BS), "bluesky_activo": fmt_date(v(ws,r,C_BS_A)),
             "mastodon": s(ws,r,C_MD), "mastodon_activo": fmt_date(v(ws,r,C_MD_A)),
             "email": s(ws,r,C_EMAIL)} for r in rows]

def build_congreso(ws, rows):
    return [{"tipo": s(ws,r,C_TIPO), "nombre": s(ws,r,C_NOMBRE), "grupo": s(ws,r,C_GRUPO),
             "circunscripcion": s(ws,r,C_CIRCUN),
             "twitter": s(ws,r,C_TW), "twitter_activo": fmt_date(v(ws,r,C_TW_A)),
             "bluesky": s(ws,r,C_BS), "bluesky_activo": fmt_date(v(ws,r,C_BS_A)),
             "mastodon": s(ws,r,C_MD), "mastodon_activo": fmt_date(v(ws,r,C_MD_A)),
             "email": s(ws,r,C_EMAIL)} for r in rows]

def build_senado(ws, rows):
    return [{"tipo": s(ws,r,C_TIPO), "nombre": s(ws,r,C_NOMBRE), "grupo": s(ws,r,C_GRUPO),
             "twitter": s(ws,r,C_TW), "twitter_activo": fmt_date(v(ws,r,C_TW_A)),
             "bluesky": s(ws,r,C_BS), "bluesky_activo": fmt_date(v(ws,r,C_BS_A)),
             "mastodon": s(ws,r,C_MD), "mastodon_activo": fmt_date(v(ws,r,C_MD_A)),
             "email": s(ws,r,C_EMAIL)} for r in rows]

def build_partidos(ws, rows):
    return [{"nombre": s(ws,r,C_NOMBRE), "ambito": "Nacional",
             "twitter": s(ws,r,C_TW), "twitter_activo": fmt_date(v(ws,r,C_TW_A)),
             "bluesky": s(ws,r,C_BS), "bluesky_activo": fmt_date(v(ws,r,C_BS_A)),
             "mastodon": s(ws,r,C_MD), "mastodon_activo": fmt_date(v(ws,r,C_MD_A)),
             "email": s(ws,r,C_EMAIL)} for r in rows]

def build_universidades(ws, rows):
    return [{"nombre": s(ws,r,C_NOMBRE), "tipo": s(ws,r,C_TIPO) or s(ws,r,C_DETALLE),
             "twitter": s(ws,r,C_TW), "twitter_activo": fmt_date(v(ws,r,C_TW_A)),
             "bluesky": s(ws,r,C_BS), "bluesky_activo": fmt_date(v(ws,r,C_BS_A)),
             "mastodon": s(ws,r,C_MD), "mastodon_activo": fmt_date(v(ws,r,C_MD_A)),
             "email": s(ws,r,C_EMAIL)} for r in rows]

def build_total(ws, rows_by_cat, refs):
    def ref_date(p):
        val = refs.get(p)
        try: return datetime.strptime(val[:10], "%Y-%m-%d").date() if val else None
        except: return None
    tw_ref, bs_ref, md_ref = ref_date("twitter"), ref_date("bluesky"), ref_date("mastodon")
    result = []
    for cat, rows in rows_by_cat.items():
        for r in rows:
            sub = s(ws,r,C_DETALLE) or s(ws,r,C_TIPO)
            result.append({
                "categoria": cat, "subcategoria": sub, "nombre": s(ws,r,C_NOMBRE),
                "twitter": s(ws,r,C_TW),
                "twitter_activo": is_active(fmt_date(v(ws,r,C_TW_A)), tw_ref),
                "bluesky": s(ws,r,C_BS),
                "bluesky_activo": is_active(fmt_date(v(ws,r,C_BS_A)), bs_ref),
                "mastodon": s(ws,r,C_MD),
                "mastodon_activo": is_active(fmt_date(v(ws,r,C_MD_A)), md_ref),
            })
    return result

CATEGORY_MAP = {
    "AGE":          ("age.json",          build_age),
    "Autonomías":   ("autonomias.json",   build_autonomias),
    "Gobierno":     ("gobierno.json",     build_gobierno),
    "Congreso":     ("congreso.json",     build_congreso),
    "Senado":       ("senado.json",       build_senado),
    "Partidos":     ("partidos.json",     build_partidos),
    "Universidades":("universidades.json",build_universidades),
}

def main():
    dry = "--dry-run" in sys.argv
    if dry: print("DRY-RUN: no se escribirá nada.\n")
    print(f"Cargando {XLSX}…")
    refs = load_refs()
    if refs: print(f"  Referencias: {refs}")
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb["Sheet1"]
    rows_by_cat = {cat: [] for cat in CATEGORY_MAP}
    for row in range(2, ws.max_row + 1):
        cat = ws.cell(row, C_CAT).value
        if cat in CATEGORY_MAP: rows_by_cat[cat].append(row)
    print()
    os.makedirs(DATA_DIR, exist_ok=True)
    for cat, (fname, builder) in CATEGORY_MAP.items():
        data = builder(ws, rows_by_cat[cat])
        dest = os.path.join(DATA_DIR, fname)
        print(f"  {fname}: {len(data)} entradas" + ("" if dry else f"  → {dest}"))
        if not dry:
            with open(dest, "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False, indent=2); f.write("\n")
    total = build_total(ws, rows_by_cat, refs)
    total_path = os.path.join(DATA_DIR, "total.json")
    print(f"  total.json: {len(total)} entradas" + ("" if dry else f"  → {total_path}"))
    if not dry:
        with open(total_path, "w", encoding="utf-8") as f:
            json.dump(total, f, ensure_ascii=False, indent=2); f.write("\n")
    print("\nHecho." if not dry else "\nDry-run completado.")
    wb.close()

if __name__ == "__main__":
    main()
