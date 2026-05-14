#!/usr/bin/env python3
"""
update_chamber.py — Actualiza una cámara (Congreso o Senado) tras elecciones.

Lee el XML oficial del Senado o el CSV del Congreso, detecta bajas y altas
respecto a datosfinales.xlsx y aplica los cambios directamente sobre él.
Crea copia de seguridad automática antes de modificar.

Fuentes oficiales:
  Senado:   https://www.senado.es/web/relacionesciudadanos/datosabiertos/catalogodatos/index.html
            → "Composición del Senado" → XML (openData…xml)
  Congreso: https://www.congreso.es/en/datos-abiertos
            → Diputados del Congreso → CSV

Uso:
    python3 dataset/scripts/update_chamber.py --camara senado --xml /ruta/al/openData.xml
    python3 dataset/scripts/update_chamber.py --camara congreso --csv /ruta/a/diputados.csv
    python3 dataset/scripts/update_chamber.py --camara senado --xml fichero.xml --dry-run

Tras actualizar, ejecutar:
    python3 dataset/scripts/infer_social.py --bluesky       # gratis
    python3 dataset/scripts/infer_social.py --twitter       # coste API
    python3 dataset/scripts/export.py
"""

import sys, os, json, argparse, unicodedata, re
from datetime import date
import openpyxl
try:
    import xml.etree.ElementTree as ET
except ImportError:
    ET = None
from backup import backup_datosfinales

SCRIPT_DIR  = os.path.dirname(os.path.abspath(__file__))
DATASET_DIR = os.path.dirname(SCRIPT_DIR)
DF_FILE     = os.path.join(DATASET_DIR, "datosfinales.xlsx")

# Columnas datosfinales (1-indexed)
C_CAT=1; C_NOMBRE=2; C_TW=3; C_TW_A=4; C_BS=5; C_BS_A=6
C_MD=7; C_MD_A=8; C_EMAIL=9; C_DETALLE=10; C_TIPO=11
C_GRUPO=12; C_CIRCUN=13; C_CCAA=14; C_PARTIDO=15

# Mapeo partido político → grupo parlamentario (Senado XV legislatura)
PARTIDO_A_GRUPO_SENADO = {
    "PP": "GPP", "PARTIDO POPULAR": "GPP",
    "PSOE": "GPS", "PARTIDO SOCIALISTA OBRERO ESPAÑOL": "GPS",
    "VOX": "GPV",
    "EAJ-PNV": "GPEAJPNV", "PARTIDO NACIONALISTA VASCO": "GPEAJPNV",
    "EH BILDU": "GPERB",
    "JXCAT": "GPPLU", "JUNTS PER CATALUNYA": "GPPLU",
    "COALICIÓN CANARIA": "GPMX", "CC": "GPMX",
    "UPN": "GPMX",
    "GEROA BAI": "GPIC", "MÉSCOMPROMÍS": "GPIC", "MÁS MADRID": "GPIC",
    "SUMAR": "GPIC", "IU": "GPIC",
    "ASG": "GPIC",
}

def norm(s):
    s = unicodedata.normalize('NFD', s.upper())
    return re.sub(r'[^A-Z ]', '', ''.join(c for c in s if unicodedata.category(c) != 'Mn')).strip()

def nombre_key(nombre):
    """Normaliza 'Apellidos, Nombre' para comparación robusta."""
    return norm(nombre)

def parse_senado_xml(xml_path):
    """Devuelve lista de dicts con datos del Senado desde el XML oficial."""
    tree = ET.parse(xml_path)
    root = tree.getroot()
    senadores = []
    for s in root.iter('senador'):
        ap = (s.findtext('apellidos') or '').strip()
        nm = (s.findtext('nombre') or '').strip()
        if not ap: continue
        nombre = f"{ap}, {nm}" if nm else ap
        partido = norm(s.findtext('partido_politico') or s.findtext('grupo_parlamentario') or '')
        grupo = PARTIDO_A_GRUPO_SENADO.get(partido, 'GPIC')
        circun = (s.findtext('circunscripcion') or '').strip()
        ccaa   = (s.findtext('comunidad_autonoma') or '').strip()
        tipo   = (s.findtext('procedencia') or '').strip()
        senadores.append({
            'nombre': nombre, 'grupo': grupo,
            'circunscripcion': circun, 'ccaa': ccaa, 'tipo': tipo,
            'partido': partido,
        })
    return senadores

def parse_congreso_csv(csv_path):
    """Devuelve lista de dicts desde CSV oficial del Congreso."""
    import csv
    diputados = []
    with open(csv_path, encoding='utf-8-sig', newline='') as f:
        reader = csv.DictReader(f, delimiter=';')
        for row in reader:
            ap = (row.get('apellidos') or row.get('Apellidos') or '').strip()
            nm = (row.get('nombre') or row.get('Nombre') or '').strip()
            if not ap: continue
            nombre = f"{ap}, {nm}" if nm else ap
            grupo = (row.get('grupo') or row.get('Grupo') or '').strip()
            circun = (row.get('circunscripcion') or row.get('Circunscripción') or '').strip()
            diputados.append({'nombre': nombre, 'grupo': grupo,
                              'circunscripcion': circun, 'partido': ''})
    return diputados

def load_df_camara(camara):
    """Lee las entradas actuales de la cámara en datosfinales."""
    wb = openpyxl.load_workbook(DF_FILE)
    ws = wb.active
    entries = {}
    for r in ws.iter_rows(min_row=2):
        cat = r[C_CAT-1].value or ''
        nombre = r[C_NOMBRE-1].value or ''
        if cat.strip() == camara and nombre:
            entries[nombre_key(nombre)] = {
                'nombre': nombre, 'row': r[0].row,
                'tw': r[C_TW-1].value, 'bs': r[C_BS-1].value, 'md': r[C_MD-1].value,
                'grupo': r[C_GRUPO-1].value,
            }
    return entries, wb, ws

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--camara', required=True, choices=['senado','congreso'])
    ap.add_argument('--xml', help='Ruta al XML oficial del Senado')
    ap.add_argument('--csv', help='Ruta al CSV oficial del Congreso')
    ap.add_argument('--dry-run', action='store_true')
    args = ap.parse_args()

    camara_label = 'Senado' if args.camara == 'senado' else 'Congreso'

    # Cargar datos nuevos
    if args.camara == 'senado':
        if not args.xml:
            print("ERROR: --xml requerido para el Senado"); sys.exit(1)
        nuevos_datos = parse_senado_xml(args.xml)
    else:
        if not args.csv:
            print("ERROR: --csv requerido para el Congreso"); sys.exit(1)
        nuevos_datos = parse_congreso_csv(args.csv)

    nuevos_keys = {nombre_key(d['nombre']): d for d in nuevos_datos}
    print(f"XML/CSV: {len(nuevos_datos)} entradas en {camara_label}")

    # Cargar estado actual en datosfinales
    actuales, wb_cr, ws_cr = load_df_camara(camara_label)
    print(f"CR actual: {len(actuales)} entradas en {camara_label}")

    # Detectar bajas (en CR pero no en nuevo)
    bajas = {k: v for k, v in actuales.items() if k not in nuevos_keys}
    # Detectar altas (en nuevo pero no en CR)
    altas = {k: v for k, v in nuevos_keys.items() if k not in actuales}
    # Detectar cambios de grupo
    cambios_grupo = {}
    for k, v in actuales.items():
        if k in nuevos_keys:
            nuevo_grupo = nuevos_keys[k].get('grupo','')
            actual_grupo = v.get('grupo','')
            if nuevo_grupo and actual_grupo and nuevo_grupo != actual_grupo:
                cambios_grupo[k] = (actual_grupo, nuevo_grupo, v['nombre'])

    print(f"\n  Bajas detectadas: {len(bajas)}")
    for k, v in bajas.items():
        print(f"    ✗ {v['nombre']}")

    print(f"\n  Altas detectadas: {len(altas)}")
    for k, v in altas.items():
        print(f"    + {v['nombre']}  ({v.get('grupo','')})")

    print(f"\n  Cambios de grupo: {len(cambios_grupo)}")
    for k, (ant, nue, nombre) in cambios_grupo.items():
        print(f"    ~ {nombre}: {ant} → {nue}")

    if args.dry_run:
        print("\nDRY-RUN: no se modifican archivos.")
        return

    if not bajas and not altas and not cambios_grupo:
        print("\nNo hay cambios que aplicar.")
        return

    print("\n¿Aplicar estos cambios? [s/N] ", end="", flush=True)
    resp = input().strip().lower()
    if resp not in ('s', 'si', 'sí', 'y', 'yes'):
        print("Cancelado.")
        return

    backup_datosfinales()

    # Aplicar en datosfinales
    # wb_cr/ws_cr ya apuntan a datosfinales (cargados en load_df_camara)
    wb_df = wb_cr; ws_df = ws_cr

    # 1. Bajas: reconstruir workbook sin esas filas (evita bug de openpyxl delete_rows)
    if bajas:
        bajas_rows = {v['row'] for v in bajas.values()}
        wb_df2 = openpyxl.Workbook()
        ws_df2 = wb_df2.active
        ws_df2.title = ws_df.title
        new_row = 1
        for r in ws_df.iter_rows(values_only=False):
            if r[0].row in bajas_rows: continue
            for j, cell in enumerate(r, 1):
                ws_df2.cell(new_row, j).value = cell.value
            new_row += 1
        wb_df = wb_df2; ws_df = ws_df2

    # 2. Cambios de grupo
    for k, (ant, nue, nombre) in cambios_grupo.items():
        for r in ws_df.iter_rows(min_row=2):
            if (r[C_CAT-1].value or '').strip() == camara_label and (r[C_NOMBRE-1].value or '').strip() == nombre:
                r[C_GRUPO-1].value = nue
                break

    # 3. Altas: añadir al final de la sección de la cámara
    last_row = max(
        (r[0].row for r in ws_df.iter_rows(min_row=2) if (r[C_CAT-1].value or '').strip() == camara_label),
        default=1
    )
    insert_at = last_row + 1
    for k, data in altas.items():
        ws_df.cell(insert_at, C_CAT).value    = camara_label
        ws_df.cell(insert_at, C_NOMBRE).value = data['nombre']
        ws_df.cell(insert_at, C_GRUPO).value  = data.get('grupo', '')
        ws_df.cell(insert_at, C_CIRCUN).value = data.get('circunscripcion', '')
        ws_df.cell(insert_at, C_CCAA).value   = data.get('ccaa', '')
        # Twitter, Bluesky, Mastodon vacíos — se rellenarán con infer_social
        insert_at += 1

    wb_df.save(DF_FILE)
    print(f"datosfinales.xlsx actualizado.")

    print(f"""
Siguiente paso recomendado:
  1. python3 dataset/scripts/infer_social.py --bluesky        # gratis
  2. python3 dataset/scripts/infer_social.py --twitter        # ⚠️ coste API
  3. python3 dataset/scripts/export.py
""")

if __name__ == "__main__":
    main()
