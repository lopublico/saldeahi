#!/usr/bin/env python3
"""
check_activity.py — Comprueba la última actividad en Bluesky y Mastodon.
Twitter requiere GetXAPI (de pago); se activa solo con --twitter.

Uso:
    python3 dataset/scripts/check_activity.py --bluesky
    python3 dataset/scripts/check_activity.py --mastodon
    python3 dataset/scripts/check_activity.py --twitter --token TU_TOKEN
    python3 dataset/scripts/check_activity.py --bluesky --mastodon   # ambos gratis

Actualiza las columnas *_Activo en datosfinales.xlsx y guarda lastUpdate.json.
"""

import sys, os, json, time, argparse, socket
from datetime import date, datetime
import urllib.request, urllib.parse, urllib.error
import openpyxl
from backup import backup_datosfinales

socket.setdefaulttimeout(15)

SCRIPT_DIR  = os.path.dirname(os.path.abspath(__file__))
DATASET_DIR = os.path.dirname(SCRIPT_DIR)
ROOT_DIR    = os.path.dirname(DATASET_DIR)
DF_FILE     = os.path.join(DATASET_DIR, "datosfinales.xlsx")
DATA_DIR    = os.path.join(ROOT_DIR, "src", "data")
LAST_UPDATE = os.path.join(DATA_DIR, "lastUpdate.json")

# Progreso de reintentos de Twitter/X: vive fuera del repo (solo en el runner)
# para poder retomar entre reintentos del mismo día sin publicar nada parcial.
TW_RESUME = os.path.join(os.environ.get('RUNNER_TEMP', '/tmp'), 'tw_resume.json')

# Columnas datosfinales (1-indexed)
C_NOMBRE=2; C_TW=3; C_TW_A=4; C_BS=5; C_BS_A=6; C_MD=7; C_MD_A=8
C_TW_CONSULTA=17; C_TW_ESTADO=18
C_BS_CONSULTA=19; C_BS_ESTADO=20
C_MD_CONSULTA=21; C_MD_ESTADO=22

GETXAPI_TOKEN = os.environ.get("GETXAPI_TOKEN", "")
GETXAPI_BASE  = "https://api.getxapi.com/twitter"


# ── Bluesky ───────────────────────────────────────────────────────────────────

def bsky_last_post(handle):
    """Devuelve 'YYYY-MM-DD' del último post, '404' si no existe, None si error."""
    h = handle.lstrip('@').lower().strip()
    if '.' not in h: h += '.bsky.social'
    url = ('https://public.api.bsky.app/xrpc/app.bsky.feed.getAuthorFeed'
           f'?actor={urllib.parse.quote(h)}&limit=1&filter=posts_no_replies')
    try:
        with urllib.request.urlopen(url, timeout=10) as r:
            data = json.loads(r.read())
        feed = data.get('feed', [])
        if not feed: return None
        ts = feed[0]['post']['record'].get('createdAt','')
        return ts[:10] if ts else None
    except urllib.error.HTTPError as e:
        body = e.read().decode()
        if '"AccountNotFound"' in body or '"InvalidRequest"' in body:
            return '404'
        return None
    except: return None


# ── Mastodon ──────────────────────────────────────────────────────────────────

def mastodon_last_post(handle):
    """Devuelve 'YYYY-MM-DD' del último toot, '404' si no existe, None si error."""
    h = handle.lstrip('@')
    if '@' not in h: return None  # necesita formato user@instance
    user, instance = h.rsplit('@', 1)
    url = f'https://{instance}/api/v1/accounts/lookup?acct={urllib.parse.quote(user)}'
    try:
        with urllib.request.urlopen(url, timeout=10) as r:
            account = json.loads(r.read())
        account_id = account.get('id')
        if not account_id: return '404'
        url2 = f'https://{instance}/api/v1/accounts/{account_id}/statuses?limit=1&exclude_replies=true'
        with urllib.request.urlopen(url2, timeout=10) as r:
            statuses = json.loads(r.read())
        if not statuses: return None
        ts = statuses[0].get('created_at', '')
        return ts[:10] if ts else None
    except urllib.error.HTTPError as e:
        if e.code == 404: return '404'
        return None
    except: return None


# ── Twitter / GetXAPI ─────────────────────────────────────────────────────────

def parse_twitter_date(ts):
    """Convierte fecha de Twitter a YYYY-MM-DD. Soporta ISO y formato nativo."""
    if not ts:
        return None
    if ts[0].isdigit():
        return ts[:10]
    try:
        from email.utils import parsedate
        t = parsedate(ts)
        if t:
            return f"{t[0]:04d}-{t[1]:02d}-{t[2]:02d}"
    except Exception:
        pass
    return None

def load_tw_resume():
    try:
        with open(TW_RESUME, encoding='utf-8') as f:
            return json.load(f).get('row', 2)
    except:
        return 2


def save_tw_resume(row):
    with open(TW_RESUME, 'w', encoding='utf-8') as f:
        json.dump({'row': row}, f)


def clear_tw_resume():
    try:
        os.remove(TW_RESUME)
    except FileNotFoundError:
        pass


def twitter_last_post(handle, token):
    h = handle.lstrip('@')
    # limit=2: el pinnado puede aparecer primero; con 2 tweets max() da el más reciente
    url = f'{GETXAPI_BASE}/user/tweets?userName={urllib.parse.quote(h)}&limit=2'
    req = urllib.request.Request(url, headers={'Authorization': f'Bearer {token}'})
    for attempt in range(2):
        try:
            with urllib.request.urlopen(req, timeout=10) as r:
                data = json.loads(r.read())
            tweets = (data.get('data') or {}).get('tweets') or data.get('tweets') or []
            if not tweets: return None
            # Extraer fechas de todos y quedarse con la más reciente
            # (el tweet pinnado puede aparecer primero pero ser más antiguo)
            dates = []
            for t in tweets:
                ts = t.get('createdAt') or t.get('created_at', '')
                d = parse_twitter_date(ts)
                if d and d != '404':
                    dates.append(d)
            return max(dates) if dates else None
        except urllib.error.HTTPError as e:
            if 400 <= e.code < 500: return '404'
            if attempt < 1: time.sleep(3)
        except:
            if attempt < 1: time.sleep(3)
    return None


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--bluesky',  action='store_true')
    ap.add_argument('--mastodon', action='store_true')
    ap.add_argument('--twitter',  action='store_true')
    ap.add_argument('--token',        default=GETXAPI_TOKEN, help='Token GetXAPI')
    ap.add_argument('--skip-updated', action='store_true', help='Saltar cuentas con fecha del mes actual')
    ap.add_argument('--only-missing', action='store_true', help='Solo cuentas sin fecha de actividad registrada (altas nuevas)')
    ap.add_argument('--time-limit',   type=int, default=150, help='Minutos máximos para la comprobación de Twitter antes de abortar (por defecto 150)')
    ap.add_argument('--dry-run',      action='store_true')
    args = ap.parse_args()

    if not any([args.bluesky, args.mastodon, args.twitter]):
        print("Indica al menos una plataforma: --bluesky, --mastodon, --twitter")
        sys.exit(1)

    if not args.dry_run:
        backup_datosfinales()

    if args.twitter and not args.token:
        print("ERROR: --twitter requiere --token TOKEN o variable GETXAPI_TOKEN")
        sys.exit(1)

    wb = openpyxl.load_workbook(DF_FILE)
    ws = wb.active
    today = date.today().isoformat()
    checked = {'bluesky': 0, 'mastodon': 0, 'twitter': 0}

    if args.bluesky or args.mastodon:
        for row in range(2, ws.max_row + 1):
            nombre = ws.cell(row, C_NOMBRE).value
            if not nombre: continue

            if args.bluesky:
                bsky = ws.cell(row, C_BS).value
                if bsky and not (args.only_missing and ws.cell(row, C_BS_A).value):
                    result = bsky_last_post(bsky)
                    if result is not None:
                        ws.cell(row, C_BS_CONSULTA).value = today
                        if result == '404':
                            # No pisar la última fecha de actividad conocida:
                            # es lo único que dice cuándo publicó por última
                            # vez antes de desaparecer.
                            ws.cell(row, C_BS_ESTADO).value = '404'
                        else:
                            ws.cell(row, C_BS_A).value = result
                            ws.cell(row, C_BS_ESTADO).value = None
                        checked['bluesky'] += 1
                    print(f"  BS {nombre}: {result}", flush=True)
                    time.sleep(0.2)

            if args.mastodon:
                md = ws.cell(row, C_MD).value
                if md and not (args.only_missing and ws.cell(row, C_MD_A).value):
                    result = mastodon_last_post(md)
                    if result is not None:
                        ws.cell(row, C_MD_CONSULTA).value = today
                        if result == '404':
                            ws.cell(row, C_MD_ESTADO).value = '404'
                        else:
                            ws.cell(row, C_MD_A).value = result
                            ws.cell(row, C_MD_ESTADO).value = None
                        checked['mastodon'] += 1
                    print(f"  MD {nombre}: {result}", flush=True)
                    time.sleep(0.2)

    tw_incomplete = False
    tw_hard_fail = False

    if args.twitter:
        # La comprobación de actividad usa una ventana de 30 días, así que una
        # pasada parcial publicada dejaría cuentas con fecha desactualizada y
        # provocaría falsos negativos. Por eso se exige recorrer TODAS las
        # cuentas antes de tocar lastUpdate.json o dar la comprobación por
        # buena. Si no da tiempo, se guarda el progreso (solo en el runner,
        # sin commitear) y se retoma desde la fila del cursor en el siguiente
        # reintento del mismo día (ver el bucle de reintentos del workflow).
        start_row = load_tw_resume()
        if not (2 <= start_row <= ws.max_row):
            start_row = 2
        if start_row > 2:
            print(f"  TW retomando desde la fila {start_row} (intento anterior incompleto)", flush=True)

        tw_deadline = time.monotonic() + args.time_limit * 60
        tw_none_streak = 0
        processed = 0
        resume_at = start_row

        for row in range(start_row, ws.max_row + 1):
            if time.monotonic() > tw_deadline:
                print(f"  TW límite de tiempo alcanzado ({args.time_limit} min) antes de terminar la lista completa", flush=True)
                tw_incomplete = True
                resume_at = row
                break
            resume_at = row + 1

            nombre = ws.cell(row, C_NOMBRE).value
            if not nombre: continue
            tw = ws.cell(row, C_TW).value
            if not tw: continue

            existing = ws.cell(row, C_TW_A).value
            # Saltar si ya tiene fecha de este mes (actualizada en este ciclo)
            if args.skip_updated and isinstance(existing, str) and existing.startswith(today[:7]):
                print(f"  TW {nombre}: ya actualizado ({existing}), saltando", flush=True)
                continue
            if args.only_missing and existing:
                continue

            result = twitter_last_post(tw, args.token)
            processed += 1
            if result is not None:
                # Sellar SIEMPRE la fecha de consulta al intentar, tenga o no
                # éxito: permite distinguir "comprobada y sin novedad/404" de
                # "no comprobada todavía".
                ws.cell(row, C_TW_CONSULTA).value = today
                if result == '404':
                    # No pisar la última fecha de actividad conocida: es lo
                    # único que nos dice cuándo publicó por última vez antes
                    # de desaparecer. El estado 404 se registra aparte.
                    ws.cell(row, C_TW_ESTADO).value = '404'
                else:
                    ws.cell(row, C_TW_A).value = result
                    ws.cell(row, C_TW_ESTADO).value = None
                checked['twitter'] += 1
                tw_none_streak = 0
            else:
                tw_none_streak += 1
                if tw_none_streak >= 5:
                    print(f"  TW {tw_none_streak} None consecutivos — posible throttling, pausando 60s", flush=True)
                    time.sleep(60)
                    tw_none_streak = 0
            print(f"  TW {nombre}: {result}", flush=True)
            time.sleep(0.3)

        if not tw_incomplete and processed > 0 and checked['twitter'] == 0:
            tw_hard_fail = True

    if not args.dry_run:
        wb.save(DF_FILE)

        if args.twitter:
            if tw_incomplete:
                save_tw_resume(resume_at)
            elif not tw_hard_fail:
                clear_tw_resume()

        # Actualizar lastUpdate.json
        try:
            with open(LAST_UPDATE, encoding='utf-8') as f:
                lu = json.load(f)
        except: lu = {}
        if args.bluesky:  lu['bluesky']  = today
        if args.mastodon: lu['mastodon'] = today
        if args.twitter and not tw_incomplete and not tw_hard_fail:
            lu['twitter'] = today
        os.makedirs(DATA_DIR, exist_ok=True)
        with open(LAST_UPDATE, 'w', encoding='utf-8') as f:
            json.dump(lu, f, indent=2)

    print(f"\nComprobados: {checked}", flush=True)
    if not args.dry_run:
        print(f"Guardado en {DF_FILE}", flush=True)
        print("Recuerda ejecutar export.py para actualizar los JSON de la web.", flush=True)

    if tw_hard_fail:
        print("ERROR: ninguna cuenta de Twitter/X pudo comprobarse (posible fallo de la API o token inválido); no se actualiza lastUpdate.", file=sys.stderr)
        sys.exit(1)
    if tw_incomplete:
        print(f"AVISO: comprobación de Twitter/X incompleta ({checked['twitter']} cuentas revisadas); progreso guardado para reintentar.", file=sys.stderr)
        sys.exit(2)

if __name__ == "__main__":
    main()
