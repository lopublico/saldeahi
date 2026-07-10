#!/usr/bin/env python3
"""
audit_invente.py — Audita la coherencia del dataset con el Inventario de Entes
del Sector Público (Invente/IGAE) y detecta altas y bajas en el universo estatal.

Uso:
    python3 dataset/scripts/audit_invente.py

Comprueba:
  A) que la forma jurídica (columna Detalle) de cada fila con código Invente
     coincide con el tipoEnte oficial del inventario
  B) filas AGE con forma jurídica pero sin código Invente
  C) bajas: códigos del Excel que ya no existen como entes vivos
  D) altas: entes nuevos del inventario que cumplen los criterios de inclusión
     y aún no tienen fila (las sociedades mercantiles requieren curación manual:
     solo se incluyen matrices)

No modifica nada: es de solo lectura. La API de Invente rechaza query strings,
por lo que se descarga el inventario completo y se filtra por DIR3 estatal (E…).
"""
import json
import os
import re
import sys
import unicodedata
import urllib.request

import openpyxl

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
DATASET_DIR = os.path.dirname(SCRIPT_DIR)
XLSX = os.path.join(DATASET_DIR, "datosfinales.xlsx")
API = "https://www.pap.hacienda.gob.es/Invente2/api"

FORMA_LABEL = {7: "Organismo Autónomo", 8: "Organismo Autónomo", 9: "Organismo Autónomo",
               14: "Agencia Estatal", 19: "Autoridad Independiente",
               4: "Entidad Pública Empresarial", 2: "Ente Público",
               20: "Entidad de Derecho Público", 3: "Entidad Gestora de la S.S.",
               6: "Fundación Estatal", 12: "Sociedad Mercantil Estatal",
               15: "Mutua Colaboradora con la S.S."}
FORMAS = set(FORMA_LABEL.values())
# tipos que entran automáticamente al universo (las sociedades, tipo 12, se curan a mano)
TIPOS_AUTO = {7, 8, 9, 14, 19, 4, 2, 20, 3, 6, 15}
# entes que nunca deben proponerse como alta
EXCLUIDOS = {
    "INV00001282",  # Consejo Económico y Social: entrada duplicada en el propio Invente
                    # (figura como tipo 2 y tipo 20); la fila del CES usa INV00001281
                    # y vive en la categoría «Órganos del Estado», no en AGE
}
# tipo 5 (Fondo sin personalidad jurídica, 16 entes) queda excluido del todo:
# son vehículos contables/financieros gestionados por un ministerio o entidad matriz
# (Tesoro, ICO/Axis, Puertos del Estado…) sin comunicación pública propia — mismo
# criterio que excluye a las micro-entidades instrumentales. No se listan como altas.
TIPOS_EXCLUIDOS_SIN_ALTA = {5}


def get(path):
    req = urllib.request.Request(API + path, headers={
        "User-Agent": "Mozilla/5.0 (saldeahi-dataset)", "Accept": "application/json"})
    with urllib.request.urlopen(req, timeout=60) as r:
        return json.load(r)


def norm_tokens(s):
    s = unicodedata.normalize("NFD", str(s or ""))
    s = "".join(c for c in s if unicodedata.category(c) != "Mn").lower()
    return set(t for t in re.split(r"[^a-z0-9]+", s) if len(t) > 3)


def main():
    print("Descargando inventario Invente…")
    entes = get("/EntidadesSPI")["EntidadesSPI"]
    vivos = {x["codigoInvente"]: x for x in entes}
    print(f"  {len(entes)} entes vivos (todos los ámbitos)")

    ws = openpyxl.load_workbook(XLSX, data_only=True)["Sheet1"]
    codigos_excel, problemas = {}, []
    for r in range(2, ws.max_row + 1):
        cat = ws.cell(r, 1).value
        nombre = str(ws.cell(r, 2).value or "").strip()
        detalle = str(ws.cell(r, 10).value or "").strip()
        codigo = str(ws.cell(r, 16).value or "").strip()
        if codigo.startswith("INV"):
            codigos_excel[codigo] = nombre
            ent = vivos.get(codigo)
            if not ent:
                problemas.append(f"C-BAJA   {nombre}: el ente {codigo} ya no está vivo en Invente")
            elif cat == "AGE":
                oficial = FORMA_LABEL.get(ent["tipoEnte"], f"tipo {ent['tipoEnte']}")
                if detalle != oficial:
                    problemas.append(f"A-FORMA  {nombre}: Excel dice «{detalle}», Invente dice «{oficial}»")
        elif cat == "AGE" and detalle in FORMAS:
            problemas.append(f"B-CODIGO {nombre}: forma jurídica «{detalle}» sin código Invente")

    # D) altas: entes estatales de tipos automáticos sin fila
    altas = [x for x in entes
             if (x["codigoDir3"] or "").startswith("E")
             and x["tipoEnte"] in TIPOS_AUTO
             and x["tipoEnte"] not in TIPOS_EXCLUIDOS_SIN_ALTA
             and "en Liquidación" not in x["nombre"]
             and x["codigoInvente"] not in codigos_excel
             and x["codigoInvente"] not in EXCLUIDOS]
    # excluir los que coinciden por nombre con una fila existente (renombrados, etc.)
    nombres_excel = [norm_tokens(n) for n in codigos_excel.values()]
    altas_netas = []
    for x in altas:
        t = norm_tokens(x["nombre"])
        if not any(len(t & n) / max(1, min(len(t), len(n))) >= 0.8 for n in nombres_excel if n):
            altas_netas.append(x)

    print()
    if problemas:
        print(f"PROBLEMAS ({len(problemas)}):")
        for p in problemas:
            print("  ", p)
    else:
        print("Sin discrepancias de forma jurídica, códigos ni bajas. ✓")
    print()
    if altas_netas:
        print(f"ALTAS CANDIDATAS ({len(altas_netas)}) — revisar e incorporar:")
        for x in altas_netas:
            print(f"   [{FORMA_LABEL.get(x['tipoEnte'], '?'):28s}] {x['nombre'][:70]}  {x['codigoInvente']}")
    else:
        print("Sin altas pendientes en el universo estatal. ✓")
    return 1 if problemas else 0


if __name__ == "__main__":
    sys.exit(main())
