#!/usr/bin/env python3
"""
infer_social.py — Inferencia de cuentas Bluesky y verificación de Twitter.

Fuentes:
  Bluesky:  seguidos de gpssenado, gpscongreso (+ semillas de partidos/gobierno)
  Twitter:  seguidos de GPSSenado, GPSCongreso, GPPopular, PPSenado
            + lookup individual de handles dudosos

Todos los datos crudos se guardan en dataset/data/ antes de procesar.
Salida: suggestions.json con propuestas (confianza + señales) para revisión.

Uso:
  python3 dataset/scripts/infer_social.py              # fetch + match
  python3 dataset/scripts/infer_social.py --no-fetch   # solo match (usa caché)
  python3 dataset/scripts/infer_social.py --apply      # escribe en datosfinales.xlsx
  python3 dataset/scripts/infer_social.py --verify-tw  # verifica handles dudosos vía API
"""

import argparse, json, os, re, sys, time, unicodedata
from collections import defaultdict
from pathlib import Path

import requests, openpyxl
from openpyxl.styles import Font, PatternFill

# ── Rutas ─────────────────────────────────────────────────────────────────────
ROOT       = Path(__file__).parent.parent.parent
DATA_DIR   = Path(__file__).parent.parent / "data"
DF_PATH    = Path(__file__).parent.parent / "datosfinales.xlsx"
SUGG_PATH  = DATA_DIR / "suggestions.json"

# Columnas datosfinales (1-indexed)
C_CAT=1; C_NOMBRE=2; C_TW=3; C_BS=5; C_MD=7; C_GRUPO=12

import sys as _sys
_sys.path.insert(0, str(Path(__file__).parent))
from backup import backup_datosfinales

DATA_DIR.mkdir(exist_ok=True)

# ── Credenciales ──────────────────────────────────────────────────────────────
TWITTER_BEARER = os.environ.get(
    "TWITTER_BEARER",
    "get-x-api-324f85bd9ab1fd3e79c6644b9be112f5effe48a508c82795"
)

# ── Cuentas semilla ───────────────────────────────────────────────────────────
BSKY_SEEDS = [
    "gpssenado.bsky.social",
    "gpscongreso.bsky.social",
]

TW_SEEDS = [
    "GPSSenado",
    "GPSCongreso",
    "GPPopular",
    "PPSenado",
]

# Handles de Twitter dudosos que necesitan verificación individual
TW_VERIFY_HANDLES = [
    "Juan_MaGonzalez",   # González Camacho, Juan Manuel (GPP)
    "LemusCandidato",    # Lemus Rubiales, Rafael Damián (GPS)
    "JuntosSotillo",     # Martín Martín, Juan Pablo (GPP)
    "mtpallares",        # Pallarès Piqué, Maria Teresa (GPPLU)
    "carmelasilva",      # Silva Rego, María Del Carmen (GPS)
    "MariaJosVillal4",   # Villalba Chavarría, María José (GPS)
    "pablodeolavide",    # Universidad Pablo de Olavide
]

# Administraciones sin Twitter/Bluesky confirmado
ADMIN_TO_SEARCH = [
    ("Adif", "adif"),
    ("Agencia Estatal de Administración Tributaria (AEAT)", "aeat agencia tributaria"),
    ("Centro de Investigaciones Sociológicas (CIS)", "cis sociologico"),
    ("Correos", "correos españa"),
    ("Instituto Geográfico Nacional (IGN)", "ign geografico nacional"),
    ("Ministerio de Agricultura, Pesca y Alimentación", "ministerio agricultura españa"),
    ("Ministerio de Industria y Turismo", "ministerio industria turismo españa"),
    ("Ministerio de Transportes y Movilidad Sostenible", "ministerio transportes españa"),
    ("Puertos del Estado", "puertos estado españa"),
    ("Servicio Público de Empleo Estatal (SEPE)", "sepe empleo españa"),
]

# Handles Bluesky que el matching automático asigna incorrectamente
BSKY_BLACKLIST = {
    "elvirodva.bsky.social", "pjsaez.bsky.social", "sir-coach-a-lot.bsky.social",
    "twistedbacteria.bsky.social", "crinamoreno.bsky.social", "libushe.eurosky.social",
    "antoniwan.online", "marcosteixeira.bsky.social", "psoe-senadoes.bsky.social",
    "xornalistas.bsky.social", "grupocema.bsky.social", "aitorlanda.bsky.social",
    "antidogmatist.bsky.social", "pedro-sanchez-archivo.braxuss.eu",
    "pablopereza.bsky.social", "xikoxikoxiko.bsky.social", "cansaludable.bsky.social",
    "portela-emilio.bsky.social",
}

# ── Helpers de normalización ──────────────────────────────────────────────────
_WORD_RE = re.compile(r"[a-záéíóúüñ]{3,}", re.IGNORECASE | re.UNICODE)

def norm(s):
    if not s: return ""
    s = unicodedata.normalize("NFD", str(s).lower())
    return "".join(c for c in s if unicodedata.category(c) != "Mn")

def tokens(s):
    return set(_WORD_RE.findall(norm(s)))

def parse_apellidos_nombre(nombre_excel):
    """'Apellidos, Nombre' → (set_apellidos, set_nombre)"""
    if "," not in str(nombre_excel):
        return tokens(nombre_excel), set()
    partes = str(nombre_excel).split(",", 1)
    aps = {t for t in tokens(partes[0]) if len(t) > 2}
    nms = {t for t in tokens(partes[1]) if len(t) > 2}
    return aps, nms

def match_score(nombre_excel, display_name):
    """Puntuación: todos los apellidos deben estar en el display_name."""
    aps, nms = parse_apellidos_nombre(nombre_excel)
    dtok = tokens(display_name)
    if not aps or not aps.issubset(dtok):
        return 0
    nm_match = bool(nms & dtok) if nms else len(aps) >= 2
    if not nm_match:
        return 0
    return len(aps) + len(nms & dtok)

def soft_match(nombre_excel, display_name):
    """Verificación blanda: 1 apellido largo + nombre, o 2+ apellidos."""
    aps, nms = parse_apellidos_nombre(nombre_excel)
    dtok = tokens(display_name)
    if not aps: return False
    aps_hit = aps & dtok
    if aps_hit:
        if nms and (nms & dtok): return True
        if len(aps_hit) >= 2: return True
        if any(len(a) > 5 for a in aps_hit): return True
    # Fallback substring para displaynames concatenados
    dc = norm(display_name).replace(" ", "")
    if len(dc) > 8:
        sub_aps = [a for a in aps if a in dc]
        if len(sub_aps) >= 2: return True
        if sub_aps and nms and any(n in dc for n in nms): return True
    return False

# ── API Bluesky ───────────────────────────────────────────────────────────────
def bsky_get_follows(actor, max_pages=30):
    follows, cursor = [], None
    for _ in range(max_pages):
        params = {"actor": actor, "limit": 100}
        if cursor:
            params["cursor"] = cursor
        try:
            r = requests.get(
                "https://public.api.bsky.app/xrpc/app.bsky.graph.getFollows",
                params=params, timeout=15)
            if r.status_code == 400:
                print(f"    400 – inválido: {actor}"); break
            if r.status_code != 200:
                print(f"    HTTP {r.status_code}: {actor}"); break
            data = r.json()
            batch = data.get("follows", [])
            follows.extend(batch)
            cursor = data.get("cursor")
            if not cursor or not batch: break
            time.sleep(0.35)
        except Exception as e:
            print(f"    Error {actor}: {e}"); break
    return follows

# ── API Twitter (GetXAPI) ─────────────────────────────────────────────────────
def tw_session():
    s = requests.Session()
    s.headers["Authorization"] = f"Bearer {TWITTER_BEARER}"
    return s

def tw_get_following(session, username, max_pages=40):
    follows, cursor = [], None
    for _ in range(max_pages):
        params = {"userName": username, "count": 200}
        if cursor:
            params["cursor"] = cursor
        try:
            r = session.get("https://api.getxapi.com/twitter/user/following",
                            params=params, timeout=30)
            if r.status_code != 200:
                print(f"    HTTP {r.status_code}: @{username}"); break
            data = r.json()
            if "error" in data:
                print(f"    Error @{username}: {data['error']}"); break
            batch = data.get("following", [])
            follows.extend(batch)
            if not data.get("has_more"): break
            cursor = data.get("next_cursor")
            time.sleep(0.5)
        except Exception as e:
            print(f"    Excepción @{username}: {e}"); break
    return follows

def tw_get_user_info(session, username):
    """Lookup individual de un handle: devuelve {name, description, ...} o None."""
    try:
        r = session.get("https://api.getxapi.com/twitter/user/info",
                        params={"userName": username}, timeout=20)
        if r.status_code != 200:
            return None
        resp = r.json()
        if "error" in resp:
            return None
        # La respuesta envuelve en {"status":..., "data": {...}}
        data = resp.get("data") or resp
        return {
            "name":        data.get("name", ""),
            "userName":    data.get("userName", username),
            "description": data.get("description") or "",
            "location":    data.get("location") or "",
            "followers":   data.get("followers", 0),
            "url":         data.get("url") or "",
        }
    except Exception:
        return None

# ── Lectura de datosfinales ────────────────────────────────────────────────────
def load_cr():
    wb = openpyxl.load_workbook(DF_PATH)
    ws = wb.active
    entries = []
    for row in ws.iter_rows(min_row=2):
        nombre = row[C_NOMBRE - 1].value
        if not nombre:
            continue
        entries.append({
            "row":     row[0].row,
            "cat":     row[C_CAT   - 1].value,
            "nombre":  nombre,
            "grupo":   row[C_GRUPO - 1].value,
            "twitter": row[C_TW    - 1].value,
            "bluesky": row[C_BS    - 1].value,
        })
    return entries

# ── FETCH ─────────────────────────────────────────────────────────────────────
def fetch_all(args):
    session = tw_session()

    # Bluesky follows
    bsky_cache = DATA_DIR / "bsky_follows.json"
    print("── Bluesky follows ──────────────────────────────────────────────")
    bsky_all = {}
    for seed in BSKY_SEEDS:
        print(f"  {seed} …", end=" ", flush=True)
        fl = bsky_get_follows(seed)
        bsky_all[seed] = [
            {"handle": f.get("handle", ""),
             "displayName": f.get("displayName", ""),
             "description": (f.get("description") or "")[:300]}
            for f in fl
        ]
        print(len(fl))
        time.sleep(0.5)
    with open(bsky_cache, "w", encoding="utf-8") as f:
        json.dump(bsky_all, f, ensure_ascii=False, indent=2)
    print(f"  → guardado en {bsky_cache}")

    # Twitter following
    tw_cache = DATA_DIR / "tw_following.json"
    print("\n── Twitter following ────────────────────────────────────────────")
    tw_all = {}
    for seed in TW_SEEDS:
        print(f"  @{seed} …", end=" ", flush=True)
        fl = tw_get_following(session, seed)
        tw_all[seed] = [
            {"userName": f.get("userName", ""),
             "name": f.get("name", ""),
             "description": (f.get("description") or "")[:300]}
            for f in fl
        ]
        print(len(fl))
        time.sleep(0.7)
    with open(tw_cache, "w", encoding="utf-8") as f:
        json.dump(tw_all, f, ensure_ascii=False, indent=2)
    print(f"  → guardado en {tw_cache}")

    # Verificación individual de handles dudosos
    if args.verify_tw:
        tw_verify_cache = DATA_DIR / "tw_verify.json"
        print("\n── Verificación individual Twitter ──────────────────────────────")
        tw_verify = {}
        for handle in TW_VERIFY_HANDLES:
            print(f"  @{handle} …", end=" ", flush=True)
            info = tw_get_user_info(session, handle)
            tw_verify[handle] = info
            if info:
                print(f"name='{info.get('name','?')}' bio='{(info.get('description','') or '')[:60]}…'")
            else:
                print("no encontrado / privado")
            time.sleep(0.6)
        with open(tw_verify_cache, "w", encoding="utf-8") as f:
            json.dump(tw_verify, f, ensure_ascii=False, indent=2)
        print(f"  → guardado en {tw_verify_cache}")

# ── MATCH ─────────────────────────────────────────────────────────────────────
BSKY_IN_BIO = re.compile(
    r'(?:^|[\s🔗:/@])([a-zA-Z0-9][a-zA-Z0-9.\-]{2,}\.(?:bsky\.social|social|cat|es|net|eu))',
    re.I | re.M
)

def extract_bsky_from_bio(bio):
    if not bio: return None
    for m in BSKY_IN_BIO.findall(bio):
        h = m.strip().lstrip("@").lower()
        if len(h) > 8 and "bsky" in h or any(
            ext in h for ext in [".bsky.social", ".social"]
        ):
            return h
    return None

def match_all():
    bsky_cache = DATA_DIR / "bsky_follows.json"
    tw_cache   = DATA_DIR / "tw_following.json"

    if not bsky_cache.exists() or not tw_cache.exists():
        print("ERROR: faltan archivos de caché. Ejecuta sin --no-fetch primero.")
        sys.exit(1)

    with open(bsky_cache, encoding="utf-8") as f:
        bsky_raw = json.load(f)
    with open(tw_cache, encoding="utf-8") as f:
        tw_raw = json.load(f)

    # Consolidar follows: handle → {displayName, description, followed_by[]}
    bsky_cons = {}
    for seed, follows in bsky_raw.items():
        for entry in follows:
            h = entry.get("handle", "").lower()
            if not h: continue
            if h not in bsky_cons:
                bsky_cons[h] = {**entry, "followed_by": []}
            bsky_cons[h]["followed_by"].append(seed)
    print(f"Bluesky: {len(bsky_cons)} cuentas únicas en follows")

    # Twitter: consolida por userName_lower
    tw_cons = {}
    for seed, follows in tw_raw.items():
        for entry in follows:
            h = entry.get("userName", "").lower()
            if not h: continue
            if h not in tw_cons:
                tw_cons[h] = {**entry, "followed_by": []}
            tw_cons[h]["followed_by"].append(seed)
    print(f"Twitter: {len(tw_cons)} cuentas únicas en follows")

    # Cargar entradas del CR
    entries = load_cr()
    needs_bsky = [e for e in entries if not e["bluesky"]
                  and e["cat"] in ("Congreso", "Senado", "Gobierno", "AGE",
                                   "Autonomías", "Partidos", "Universidades")]
    print(f"CR: {len(needs_bsky)} entradas sin Bluesky")

    # --- Matching Bluesky desde follows ---
    suggestions = []

    # Paso 1: por display name (score estricto)
    matched_handles = set()
    for entry in needs_bsky:
        best_h, best_sc, best_info = None, 0, None
        for h, info in bsky_cons.items():
            if h in matched_handles: continue
            if h in BSKY_BLACKLIST: continue
            sc = match_score(entry["nombre"], info.get("displayName", ""))
            if sc > best_sc:
                best_sc = sc
                best_h = h
                best_info = info

        if best_h and best_sc >= 2:
            suggestions.append({
                "nombre":    entry["nombre"],
                "cat":       entry["cat"],
                "grupo":     entry["grupo"],
                "tipo":      "bluesky_por_display",
                "handle":    best_h,
                "display":   best_info.get("displayName", ""),
                "bio":       (best_info.get("description") or "")[:200],
                "score":     best_sc,
                "seeds":     best_info.get("followed_by", []),
                "confianza": "Alta" if best_sc >= 3 else "Media",
                "tw_actual": entry["twitter"],
            })
            matched_handles.add(best_h)

    # Paso 2: por handle (apellido largo contenido en el handle)
    already_suggested = {s["nombre"] for s in suggestions}
    for entry in needs_bsky:
        if entry["nombre"] in already_suggested: continue
        aps, _ = parse_apellidos_nombre(entry["nombre"])
        long_aps = [a for a in aps if len(a) > 5]
        for h, info in bsky_cons.items():
            if h in matched_handles or h in BSKY_BLACKLIST: continue
            if any(a in h for a in long_aps):
                sc_dn = match_score(entry["nombre"], info.get("displayName", ""))
                suggestions.append({
                    "nombre":    entry["nombre"],
                    "cat":       entry["cat"],
                    "grupo":     entry["grupo"],
                    "tipo":      "bluesky_por_handle",
                    "handle":    h,
                    "display":   info.get("displayName", ""),
                    "bio":       (info.get("description") or "")[:200],
                    "score":     1.5,
                    "score_dn":  sc_dn,
                    "seeds":     info.get("followed_by", []),
                    "confianza": "Baja",
                    "tw_actual": entry["twitter"],
                })
                matched_handles.add(h)
                already_suggested.add(entry["nombre"])
                break

    # Paso 3: Bluesky extraído de la bio de Twitter (para los que tienen Twitter)
    already_suggested2 = {s["nombre"] for s in suggestions}
    for entry in needs_bsky:
        if entry["nombre"] in already_suggested2: continue
        if not entry["twitter"]: continue
        h_tw = str(entry["twitter"]).lstrip("@").lower()
        info_tw = tw_cons.get(h_tw)
        if info_tw:
            bsky_from_bio = extract_bsky_from_bio(info_tw.get("description", ""))
            if bsky_from_bio and bsky_from_bio not in BSKY_BLACKLIST:
                suggestions.append({
                    "nombre":    entry["nombre"],
                    "cat":       entry["cat"],
                    "grupo":     entry["grupo"],
                    "tipo":      "bluesky_desde_bio_twitter",
                    "handle":    bsky_from_bio,
                    "display":   info_tw.get("name", ""),
                    "bio":       (info_tw.get("description") or "")[:200],
                    "score":     3,
                    "seeds":     [f"bio_de_@{h_tw}"],
                    "confianza": "Alta",
                    "tw_actual": entry["twitter"],
                })
                already_suggested2.add(entry["nombre"])

    # --- Matching Twitter (para los que no tienen Twitter) ---
    needs_tw = [e for e in entries if not e["twitter"]
                and e["cat"] in ("Congreso", "Senado")
                and "," in str(e["nombre"])]

    matched_tw = set()
    for entry in needs_tw:
        best_h, best_sc, best_info = None, 0, None
        for h, info in tw_cons.items():
            if h in matched_tw: continue
            sc = match_score(entry["nombre"], info.get("name", ""))
            if sc > best_sc:
                best_sc = sc
                best_h = h
                best_info = info

        # También intentar soft match si no hay score estricto >=2
        if best_sc < 2:
            for h, info in tw_cons.items():
                if h in matched_tw: continue
                if soft_match(entry["nombre"], info.get("name", "")):
                    sc2 = 1.5
                    if sc2 > best_sc:
                        best_sc = sc2
                        best_h = h
                        best_info = info

        if best_h and best_sc >= 1.5:
            suggestions.append({
                "nombre":    entry["nombre"],
                "cat":       entry["cat"],
                "grupo":     entry["grupo"],
                "tipo":      "twitter_desde_following",
                "handle":    "@" + best_info.get("userName", best_h),
                "display":   best_info.get("name", ""),
                "bio":       (best_info.get("description") or "")[:200],
                "score":     best_sc,
                "seeds":     best_info.get("followed_by", []),
                "confianza": "Alta" if best_sc >= 3 else ("Media" if best_sc >= 2 else "Baja"),
                "tw_actual": None,
            })
            matched_tw.add(best_h)

    # --- Verificación de handles dudosos de Twitter ---
    tw_verify_cache = DATA_DIR / "tw_verify.json"
    if tw_verify_cache.exists():
        with open(tw_verify_cache, encoding="utf-8") as f:
            tw_verify = json.load(f)

        verify_results = []
        handle_to_nombre = {
            "Juan_MaGonzalez":  "González Camacho, Juan Manuel",
            "LemusCandidato":   "Lemus Rubiales, Rafael Damián",
            "JuntosSotillo":    "Martín Martín, Juan Pablo",
            "mtpallares":       "Pallarès Piqué, Maria Teresa",
            "carmelasilva":     "Silva Rego, María Del Carmen",
            "MariaJosVillal4":  "Villalba Chavarría, María José",
            "pablodeolavide":   "Universidad Pablo de Olavide",
        }
        for handle, nombre in handle_to_nombre.items():
            info = tw_verify.get(handle)
            if info is None:
                veredicto = "NO_EXISTE"
                display = ""
                bio = ""
            else:
                display = info.get("name") or info.get("userName", "")
                bio = (info.get("description") or "")[:200]
                sc = match_score(nombre, display)
                soft = soft_match(nombre, display)
                if sc >= 2 or soft:
                    veredicto = "OK"
                elif sc == 0 and not soft:
                    veredicto = "INCORRECTO"
                else:
                    veredicto = "DUDOSO"
            verify_results.append({
                "nombre":    nombre,
                "handle":    handle,
                "display":   display,
                "bio":       bio,
                "veredicto": veredicto,
            })

        print("\n── Verificación handles Twitter dudosos ─────────────────────────")
        for v in verify_results:
            flag = {"OK": "✓", "NO_EXISTE": "✗ NO EXISTE", "INCORRECTO": "✗ INCORRECTO", "DUDOSO": "⚠"}.get(v["veredicto"], "?")
            print(f"  {flag} @{v['handle']} → '{v['display']}'")
            if v["bio"]:
                print(f"       bio: {v['bio'][:80]}")
    else:
        verify_results = []

    # Guardar sugerencias
    output = {
        "bluesky_sugerencias": [s for s in suggestions if "bluesky" in s["tipo"]],
        "twitter_sugerencias": [s for s in suggestions if "twitter" in s["tipo"]],
        "tw_verificacion":     verify_results,
    }
    with open(SUGG_PATH, "w", encoding="utf-8") as f:
        json.dump(output, f, ensure_ascii=False, indent=2)

    print(f"\n── Resumen ──────────────────────────────────────────────────────")
    bs_alt  = sum(1 for s in output["bluesky_sugerencias"] if s["confianza"] == "Alta")
    bs_med  = sum(1 for s in output["bluesky_sugerencias"] if s["confianza"] == "Media")
    bs_baj  = sum(1 for s in output["bluesky_sugerencias"] if s["confianza"] == "Baja")
    tw_alt  = sum(1 for s in output["twitter_sugerencias"] if s["confianza"] == "Alta")
    tw_med  = sum(1 for s in output["twitter_sugerencias"] if s["confianza"] == "Media")
    tw_baj  = sum(1 for s in output["twitter_sugerencias"] if s["confianza"] == "Baja")
    print(f"  Bluesky: {len(output['bluesky_sugerencias'])} sugerencias "
          f"(Alta={bs_alt} Media={bs_med} Baja={bs_baj})")
    print(f"  Twitter: {len(output['twitter_sugerencias'])} sugerencias "
          f"(Alta={tw_alt} Media={tw_med} Baja={tw_baj})")
    print(f"  Sugerencias guardadas en {SUGG_PATH}")

    return output

# ── APPLY ─────────────────────────────────────────────────────────────────────
def apply_suggestions(min_confianza="Alta"):
    """Escribe en datosfinales.xlsx las sugerencias de alta confianza."""
    if not SUGG_PATH.exists():
        print("ERROR: ejecuta primero sin --apply para generar suggestions.json")
        sys.exit(1)

    with open(SUGG_PATH, encoding="utf-8") as f:
        output = json.load(f)

    backup_datosfinales()

    wb = openpyxl.load_workbook(DF_PATH)
    ws = wb.active

    # Índice nombre → fila
    row_idx = {}
    for row in ws.iter_rows(min_row=2):
        n = row[C_NOMBRE - 1].value
        if n:
            row_idx[n] = row[0].row

    levels = {"Alta": 3, "Media": 2, "Baja": 1}
    min_level = levels.get(min_confianza, 3)

    applied_bs = applied_tw = 0

    for s in output["bluesky_sugerencias"]:
        if levels.get(s["confianza"], 0) < min_level: continue
        r = row_idx.get(s["nombre"])
        if not r: continue
        cell = ws.cell(r, C_BS)
        if cell.value: continue  # ya tiene Bluesky, no sobreescribir
        handle = s["handle"]
        if handle and "." not in handle:
            handle = handle + ".bsky.social"
        cell.value = handle
        applied_bs += 1

    for s in output["twitter_sugerencias"]:
        if levels.get(s["confianza"], 0) < min_level: continue
        r = row_idx.get(s["nombre"])
        if not r: continue
        cell = ws.cell(r, C_TW)
        if cell.value: continue
        cell.value = s["handle"].lstrip("@")
        applied_tw += 1

    wb.save(DF_PATH)
    print(f"Aplicados: {applied_bs} Bluesky, {applied_tw} Twitter (confianza≥{min_confianza})")

# ── Main ──────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--bluesky",    action="store_true", help="Fetch seguidos Bluesky [gratis]")
    parser.add_argument("--twitter",    action="store_true", help="Fetch seguidos Twitter [coste API]")
    parser.add_argument("--no-fetch",   action="store_true", help="Usar caché existente, no hacer fetch")
    parser.add_argument("--verify-bsky",action="store_true", help="Verificar bios/posts Bluesky de handles aplicados")
    parser.add_argument("--verify-tw",  action="store_true", help="Verificar handles Twitter dudosos via GetXAPI")
    parser.add_argument("--apply",      action="store_true", help="Aplicar sugerencias Alta/Media al CR")
    parser.add_argument("--min-conf",   default="Alta",      help="Confianza mínima al aplicar (Alta/Media/Baja)")
    args = parser.parse_args()

    if not args.no_fetch:
        fetch_all(args)

    output = match_all()

    if args.apply:
        apply_suggestions(args.min_conf)
