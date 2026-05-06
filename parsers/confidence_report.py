#!/usr/bin/env python3
"""
confidence_report.py — Genera un xlsx de revisión de cuentas X, Bluesky y Mastodon.

Uso:
    python3 parsers/confidence_report.py
    python3 parsers/confidence_report.py --output revision.xlsx
    python3 parsers/confidence_report.py --category congreso
"""

import json
import os
import re
import argparse
import unicodedata
from datetime import date, datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
ROOT_DIR   = os.path.dirname(SCRIPT_DIR)
DATA_DIR   = os.path.join(ROOT_DIR, "src", "data")
OUTPUT_DEFAULT = os.path.join(ROOT_DIR, "confidence_report.xlsx")

CATEGORY_FILES = {
    "age":           "age.json",
    "gobierno":      "gobierno.json",
    "congreso":      "congreso.json",
    "senado":        "senado.json",
    "partidos":      "partidos.json",
    "autonomias":    "autonomias.json",
    "universidades": "universidades.json",
}

CAT_LABELS = {
    "age":           "Administración",
    "gobierno":      "Gobierno",
    "congreso":      "Congreso",
    "senado":        "Senado",
    "partidos":      "Partidos",
    "autonomias":    "Autonomías",
    "universidades": "Universidades",
}


# ── Heurística de confianza Twitter ──────────────────────────────────────────

def normalize_str(s: str) -> str:
    return unicodedata.normalize("NFD", s).encode("ascii", "ignore").decode().lower()

def name_tokens(name: str) -> list[str]:
    stop = {"de", "del", "la", "el", "los", "las", "y", "e", "i", "en", "por"}
    return [t for t in re.split(r"[\s,.\-/]+", normalize_str(name)) if len(t) >= 3 and t not in stop]

def twitter_score(item: dict, nombre: str) -> tuple[int, list[str]]:
    handle = item.get("twitter")
    activo = item.get("twitter_activo")
    notes  = []

    if not handle:
        return -1, ["Sin cuenta"]

    if activo == "404":
        return 5, ["Handle 404 (no encontrado o suspendido)"]

    score = 35
    h_lower = handle.lower()

    if activo and re.match(r"^\d{4}-\d{2}-\d{2}", str(activo)):
        try:
            days_ago = (date.today() - datetime.strptime(str(activo)[:10], "%Y-%m-%d").date()).days
            if days_ago <= 30:
                score += 15; notes.append(f"Activa: {days_ago}d")
            elif days_ago <= 180:
                score += 8;  notes.append(f"Poco activa: {days_ago}d")
            else:
                notes.append(f"Inactiva: {days_ago}d")
        except ValueError:
            pass
    else:
        notes.append("Sin fecha de actividad")

    matched = [t for t in name_tokens(nombre) if t in normalize_str(handle)]
    if len(matched) >= 2:
        score += 30; notes.append(f"Handle: tokens '{', '.join(matched)}'")
    elif len(matched) == 1:
        score += 18; notes.append(f"Handle: token '{matched[0]}'")
    else:
        initials = "".join(t[0] for t in name_tokens(nombre))
        if len(initials) >= 3 and initials[:3] in h_lower:
            score += 10; notes.append(f"Posibles siglas ({initials[:4]})")
        else:
            notes.append("Sin coincidencia de nombre — revisar")

    if any(kw in h_lower for kw in ["gob", "gov", "spain", "es_", "_es", "official", "oficial"]):
        score += 10; notes.append("Patrón institucional")

    if len(handle) <= 4:
        score -= 15; notes.append("Handle muy corto (≤4)")
    elif len(handle) <= 6 and not matched:
        score -= 5;  notes.append("Handle corto sin match")

    if re.search(r"_{2,}|\d{4,}", handle):
        score -= 5; notes.append("Números o doble guión")

    return min(max(score, 0), 100), notes


def score_label(score: int) -> str:
    if score < 0:   return "—"
    if score >= 75: return "Alta"
    if score >= 50: return "Media"
    if score >= 25: return "Baja"
    return "Muy baja"


# ── Carga de datos ────────────────────────────────────────────────────────────

def load_all(only_category=None) -> list[dict]:
    entries = []
    for cat, fname in CATEGORY_FILES.items():
        if only_category and cat != only_category:
            continue
        path = os.path.join(DATA_DIR, fname)
        with open(path, encoding="utf-8") as f:
            for item in json.load(f):
                entries.append({"category": cat, "item": item})
    return entries


# ── Estilos ───────────────────────────────────────────────────────────────────

HEADER_FILL = PatternFill("solid", fgColor="1E293B")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)

SCORE_FILLS = {
    "Alta":     PatternFill("solid", fgColor="D1FAE5"),
    "Media":    PatternFill("solid", fgColor="FEF9C3"),
    "Baja":     PatternFill("solid", fgColor="FFEDD5"),
    "Muy baja": PatternFill("solid", fgColor="FEE2E2"),
    "—":        PatternFill("solid", fgColor="F1F5F9"),
}

CAT_FILLS = {
    "age":           PatternFill("solid", fgColor="EDE9FE"),
    "gobierno":      PatternFill("solid", fgColor="DBEAFE"),
    "congreso":      PatternFill("solid", fgColor="D1FAE5"),
    "senado":        PatternFill("solid", fgColor="FEF3C7"),
    "partidos":      PatternFill("solid", fgColor="FCE7F3"),
    "autonomias":    PatternFill("solid", fgColor="CCFBF1"),
    "universidades": PatternFill("solid", fgColor="FEE2E2"),
}

THIN = Side(style="thin", color="E2E8F0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

LINK_FONT = Font(color="2563EB", underline="single", size=11)


def write_header(ws, cols: list[str]):
    for i, label in enumerate(cols, 1):
        cell = ws.cell(row=1, column=i, value=label)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    ws.row_dimensions[1].height = 28


def style_cell(cell, fill=None, bold=False, wrap=False, align="left"):
    cell.border = BORDER
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    if fill:
        cell.fill = fill
    if bold:
        cell.font = Font(bold=True, size=11)


# ── Generación xlsx ───────────────────────────────────────────────────────────

COLUMNS = [
    ("Categoría",        12),
    ("Nombre",           40),
    ("X / Twitter",      22),
    ("Última actividad", 16),
    ("Confianza",        12),
    ("Señales",          42),
    ("Bluesky",          30),
    ("Mastodon",         35),
]


def build_xlsx(entries: list[dict], output_path: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "Revisión de cuentas"

    write_header(ws, [c[0] for c in COLUMNS])

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"

    for col_idx, (_, width) in enumerate(COLUMNS, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    for row_idx, entry in enumerate(entries, start=2):
        cat  = entry["category"]
        item = entry["item"]
        nombre = item.get("nombre", "")

        tw_handle = item.get("twitter")
        tw_activo = item.get("twitter_activo")
        tw_score, tw_notes = twitter_score(item, nombre)
        tw_label = score_label(tw_score)

        bluesky  = item.get("bluesky")
        mastodon = item.get("mastodon")

        cat_fill   = CAT_FILLS.get(cat)
        score_fill = SCORE_FILLS.get(tw_label, PatternFill())

        # Col 1 — Categoría
        c1 = ws.cell(row=row_idx, column=1, value=CAT_LABELS.get(cat, cat))
        style_cell(c1, fill=cat_fill, align="center")

        # Col 2 — Nombre
        c2 = ws.cell(row=row_idx, column=2, value=nombre)
        style_cell(c2, bold=True)

        # Col 3 — X/Twitter (link)
        if tw_handle:
            url = f"https://x.com/{tw_handle}"
            c3 = ws.cell(row=row_idx, column=3, value=f"@{tw_handle}")
            c3.hyperlink = url
            c3.font = LINK_FONT
            c3.border = BORDER
            c3.alignment = Alignment(horizontal="left", vertical="center")
            if tw_activo == "404":
                c3.font = Font(color="DC2626", underline="single", size=11)
        else:
            c3 = ws.cell(row=row_idx, column=3, value="—")
            style_cell(c3, align="center")

        # Col 4 — Última actividad
        c4 = ws.cell(row=row_idx, column=4, value=str(tw_activo) if tw_activo and tw_activo != "404" else ("404" if tw_activo == "404" else ""))
        style_cell(c4, align="center")
        if tw_activo == "404":
            c4.font = Font(color="DC2626", bold=True, size=11)

        # Col 5 — Confianza
        c5 = ws.cell(row=row_idx, column=5, value=f"{tw_label}" + (f" ({tw_score})" if tw_score >= 0 else ""))
        style_cell(c5, fill=score_fill, align="center")

        # Col 6 — Señales
        c6 = ws.cell(row=row_idx, column=6, value=" · ".join(tw_notes))
        style_cell(c6, wrap=True)

        # Col 7 — Bluesky (link)
        if bluesky:
            url = f"https://bsky.app/profile/{bluesky}"
            c7 = ws.cell(row=row_idx, column=7, value=bluesky)
            c7.hyperlink = url
            c7.font = LINK_FONT
            c7.border = BORDER
            c7.alignment = Alignment(horizontal="left", vertical="center")
        else:
            c7 = ws.cell(row=row_idx, column=7, value="—")
            style_cell(c7, align="center")

        # Col 8 — Mastodon (link)
        if mastodon:
            c8 = ws.cell(row=row_idx, column=8, value=mastodon)
            c8.hyperlink = mastodon
            c8.font = LINK_FONT
            c8.border = BORDER
            c8.alignment = Alignment(horizontal="left", vertical="center")
        else:
            c8 = ws.cell(row=row_idx, column=8, value="—")
            style_cell(c8, align="center")

        ws.row_dimensions[row_idx].height = 20

    wb.save(output_path)
    print(f"✅ Guardado en: {output_path}  ({len(entries)} filas)")


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Genera xlsx de revisión de cuentas")
    parser.add_argument("--output", "-o", default=OUTPUT_DEFAULT)
    parser.add_argument("--category", "-c", choices=list(CATEGORY_FILES.keys()), default=None)
    args = parser.parse_args()

    entries = load_all(args.category)
    print(f"Cargadas {len(entries)} entidades")
    build_xlsx(entries, args.output)


if __name__ == "__main__":
    main()
