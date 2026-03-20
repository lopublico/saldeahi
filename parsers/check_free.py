#!/usr/bin/env python3
"""
check_free.py — Actualiza las fechas de último post de Bluesky y Mastodon.

Sin coste, sin límites de peticiones.

Uso:
    python3 check_free.py              # comprueba Bluesky + Mastodon
    python3 check_free.py --bluesky    # solo Bluesky
    python3 check_free.py --mastodon   # solo Mastodon

Resultado en la celda "Activo":
    "YYYY-MM-DD"  → fecha del último post
    "404"         → handle no encontrado
    vacío         → sin handle o sin posts visibles
"""

import json
import os
import re
import sys
import time
import requests
import openpyxl
from datetime import datetime, timezone

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
ROOT_DIR   = os.path.dirname(SCRIPT_DIR)
XLSX       = os.path.join(ROOT_DIR, "datosfinales.xlsx")
BSKY_API   = "https://public.api.bsky.app/xrpc"
SAVE_EVERY = 25   # guardar cada N filas procesadas
BSKY_DELAY = 0.4  # segundos entre peticiones (evita feeds vacíos por rate-limit)

# Columnas (1-indexed)
COL_BS   = 5
COL_BS_A = 6
COL_MD   = 7
COL_MD_A = 8


def fmt(dt: datetime) -> str:
    return dt.strftime("%Y-%m-%d")


def is_date(val) -> bool:
    """True si la celda ya tiene una fecha válida almacenada."""
    if isinstance(val, datetime):
        return True
    return isinstance(val, str) and bool(re.match(r"^\d{4}-\d{2}-\d{2}", val))


# ── Bluesky ───────────────────────────────────────────────────────────────────

def bsky_last_post(handle: str):
    """Devuelve 'YYYY-MM-DD', '404' o None (sin posts / error)."""
    try:
        r = requests.get(
            f"{BSKY_API}/app.bsky.feed.getAuthorFeed",
            params={"actor": handle, "limit": 1},
            timeout=8,
        )
        if r.status_code == 400:
            # Profile not found o handle inválido
            return "404"
        if r.status_code != 200:
            return None
        feed = r.json().get("feed", [])
        if not feed:
            return None   # perfil existe pero sin posts públicos
        ts = feed[0]["post"]["indexedAt"]
        dt = datetime.fromisoformat(ts.replace("Z", "+00:00"))
        return fmt(dt)
    except Exception:
        return None


def check_bluesky(ws, wb):
    print("Comprobando Bluesky…")
    checked = updated = not_found = 0

    for row in range(2, ws.max_row + 1):
        handle = ws.cell(row, COL_BS).value
        if not handle:
            ws.cell(row, COL_BS_A).value = None
            continue

        result = bsky_last_post(handle)
        current = ws.cell(row, COL_BS_A).value

        if result == "404":
            ws.cell(row, COL_BS_A).value = "404"
            not_found += 1
        elif result:
            ws.cell(row, COL_BS_A).value = result
            updated += 1
        elif not is_date(current):
            # Sin resultado y sin fecha previa → dejar vacío
            ws.cell(row, COL_BS_A).value = None
        # Si result=None pero ya hay una fecha guardada, la conservamos

        checked += 1
        time.sleep(BSKY_DELAY)

        if checked % SAVE_EVERY == 0:
            wb.save(XLSX)
            print(f"  {checked} procesados…")

    wb.save(XLSX)
    print(f"  Bluesky: {checked} comprobados — {updated} con fecha, {not_found} no encontrados.\n")


# ── Mastodon ──────────────────────────────────────────────────────────────────

def mastodon_last_post(handle: str):
    """
    handle formato: https://instancia.tld/@usuario
    Devuelve 'YYYY-MM-DD', '404' o None.
    """
    from urllib.parse import urlparse
    parsed = urlparse(handle)
    if not parsed.netloc or not parsed.path.startswith('/@'):
        return None
    instance = parsed.netloc
    user = parsed.path.lstrip('/@')
    try:
        r = requests.get(
            f"https://{instance}/api/v1/accounts/lookup",
            params={"acct": user},
            timeout=8,
        )
        if r.status_code == 404:
            return "404"
        if r.status_code != 200:
            return None
        acct_id = r.json().get("id")
        if not acct_id:
            return None

        r2 = requests.get(
            f"https://{instance}/api/v1/accounts/{acct_id}/statuses",
            params={"limit": 1, "exclude_replies": "true"},
            timeout=8,
        )
        if r2.status_code != 200:
            return None
        statuses = r2.json()
        if not statuses:
            return None
        dt = datetime.fromisoformat(statuses[0]["created_at"].replace("Z", "+00:00"))
        return fmt(dt)
    except Exception:
        return None


def check_mastodon(ws, wb):
    print("Comprobando Mastodon…")
    checked = updated = not_found = 0

    for row in range(2, ws.max_row + 1):
        handle = ws.cell(row, COL_MD).value
        if not handle:
            ws.cell(row, COL_MD_A).value = None
            continue

        result = mastodon_last_post(handle)
        current = ws.cell(row, COL_MD_A).value

        if result == "404":
            ws.cell(row, COL_MD_A).value = "404"
            not_found += 1
        elif result:
            ws.cell(row, COL_MD_A).value = result
            updated += 1
        elif not is_date(current):
            ws.cell(row, COL_MD_A).value = None

        checked += 1

        if checked % SAVE_EVERY == 0:
            wb.save(XLSX)

    wb.save(XLSX)
    if checked:
        print(f"  Mastodon: {checked} comprobados — {updated} con fecha, {not_found} no encontrados.\n")
    else:
        print("  Mastodon: no hay handles en el archivo.\n")


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    args = set(sys.argv[1:])
    only_bs = "--bluesky"  in args
    only_md = "--mastodon" in args
    run_all = not (only_bs or only_md)

    print(f"Cargando {XLSX}…")
    wb = openpyxl.load_workbook(XLSX)
    ws = wb["Sheet1"]

    try:
        if run_all or only_bs:
            check_bluesky(ws, wb)

        if run_all or only_md:
            check_mastodon(ws, wb)
    finally:
        last_update_path = os.path.join(ROOT_DIR, "src", "data", "lastUpdate.json")
        try:
            with open(last_update_path, encoding="utf-8") as f:
                data = json.load(f)
        except (FileNotFoundError, json.JSONDecodeError):
            data = {}
        today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
        if run_all or only_bs:
            data["bluesky"] = today
        if run_all or only_md:
            data["mastodon"] = today
        with open(last_update_path, "w", encoding="utf-8") as f:
            json.dump(data, f)
        print(f"Fecha de última actualización guardada en {last_update_path}")

    print("Hecho.")


if __name__ == "__main__":
    main()
