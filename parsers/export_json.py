#!/usr/bin/env python3
"""
export_json.py — Genera los archivos JSON de src/data/ a partir de datosfinales.xlsx.

Uso:
    python3 parsers/export_json.py             # exporta todos los archivos
    python3 parsers/export_json.py --dry-run   # muestra resumen sin escribir nada

Archivos generados:
    src/data/age.json
    src/data/autonomias.json
    src/data/gobierno.json
    src/data/congreso.json
    src/data/senado.json
    src/data/partidos.json
    src/data/universidades.json
    src/data/total.json

Columnas del Excel (1-indexed):
    1  Categoría       → determina el archivo JSON de destino
    2  Nombre
    3  Twitter         → handle
    4  Twitter Activo  → "YYYY-MM-DD" | "404" | vacío
    5  Bluesky         → handle
    6  Bluesky Activo  → "YYYY-MM-DD" | "404" | vacío
    7  Mastodon        → handle
    8  Mastodon Activo → "YYYY-MM-DD" | "404" | vacío
    9  Email
   10  Detalle         → categoria (AGE) | cargo (Gobierno)
   11  Tipo            → tipo (Autonomías, Congreso, Senado, Universidades)
   12  Grupo           → grupo parlamentario (Congreso, Senado)
   13  Circunscripción → circunscripcion (Congreso)
   14  CCAA            → comunidad autónoma (Autonomías)
   15  Partido         → partido del presidente (Autonomías)
"""

import re
import sys
import json
import os
from datetime import datetime, timedelta, date
from typing import Optional
import openpyxl

# ── Rutas ─────────────────────────────────────────────────────────────────────

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
ROOT_DIR   = os.path.dirname(SCRIPT_DIR)
XLSX       = os.path.join(ROOT_DIR, "datosfinales.xlsx")
DATA_DIR   = os.path.join(ROOT_DIR, "src", "data")

# ── Columnas (1-indexed) ──────────────────────────────────────────────────────

C_CAT     = 1
C_NOMBRE  = 2
C_TW      = 3
C_TW_A    = 4
C_BS      = 5
C_BS_A    = 6
C_MD      = 7
C_MD_A    = 8
C_EMAIL   = 9
C_DETALLE = 10
C_TIPO    = 11
C_GRUPO   = 12
C_CIRCUN  = 13
C_CCAA    = 14
C_PARTIDO = 15

# ── Utilidades ────────────────────────────────────────────────────────────────

def fmt_date(val):
    """
    Normaliza un valor de celda de fecha:
      - datetime  → "YYYY-MM-DD"
      - "404"     → "404"
      - "YYYY-MM-DD..." → "YYYY-MM-DD" (primeros 10 caracteres)
      - None / vacío → None
    """
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    s = str(val).strip()
    if not s:
        return None
    if s == "404":
        return "404"
    if re.match(r"^\d{4}-\d{2}-\d{2}", s):
        return s[:10]
    return None


def load_last_update_refs() -> dict:
    """Lee lastUpdate.json y devuelve las fechas de referencia por plataforma."""
    path = os.path.join(DATA_DIR, "lastUpdate.json")
    try:
        with open(path, encoding="utf-8") as f:
            return json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        return {}


def is_active(val, ref_date: Optional[date] = None) -> bool:
    """True si val es una fecha dentro de los 30 días anteriores a ref_date.

    ref_date debe ser la fecha del último check, no la fecha actual.
    Si no se pasa ref_date, se usa hoy como fallback.
    """
    if not isinstance(val, str):
        return False
    m = re.match(r"^(\d{4}-\d{2}-\d{2})", val)
    if not m:
        return False
    try:
        d = datetime.strptime(m.group(1), "%Y-%m-%d").date()
        ref = ref_date if ref_date else datetime.now().date()
        cutoff = ref - timedelta(days=30)
        return cutoff <= d <= ref
    except ValueError:
        return False


def v(ws, row: int, col: int):
    """Lee el valor de una celda; devuelve None si está vacía."""
    val = ws.cell(row, col).value
    if val == "" or val is None:
        return None
    return val


def str_val(ws, row: int, col: int):
    """Devuelve el valor como string sin espacios, o None."""
    val = v(ws, row, col)
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None


# ── Constructores por categoría ───────────────────────────────────────────────

def build_age(ws, rows):
    """AGE → age.json"""
    result = []
    for row in rows:
        result.append({
            "nombre":          str_val(ws, row, C_NOMBRE),
            "categoria":       str_val(ws, row, C_DETALLE),
            "twitter":         str_val(ws, row, C_TW),
            "twitter_activo":  fmt_date(v(ws, row, C_TW_A)),
            "bluesky":         str_val(ws, row, C_BS),
            "bluesky_activo":  fmt_date(v(ws, row, C_BS_A)),
            "mastodon":        str_val(ws, row, C_MD),
            "mastodon_activo": fmt_date(v(ws, row, C_MD_A)),
            "email":           str_val(ws, row, C_EMAIL),
        })
    return result


def build_autonomias(ws, rows):
    """Autonomías → autonomias.json"""
    result = []
    for row in rows:
        result.append({
            "tipo":            str_val(ws, row, C_TIPO),
            "ccaa":            str_val(ws, row, C_CCAA),
            "nombre":          str_val(ws, row, C_NOMBRE),
            "partido":         str_val(ws, row, C_PARTIDO),
            "twitter":         str_val(ws, row, C_TW),
            "twitter_activo":  fmt_date(v(ws, row, C_TW_A)),
            "bluesky":         str_val(ws, row, C_BS),
            "bluesky_activo":  fmt_date(v(ws, row, C_BS_A)),
            "mastodon":        str_val(ws, row, C_MD),
            "mastodon_activo": fmt_date(v(ws, row, C_MD_A)),
            "email":           str_val(ws, row, C_EMAIL),
        })
    return result


def build_gobierno(ws, rows):
    """Gobierno → gobierno.json"""
    result = []
    for row in rows:
        result.append({
            "nombre":          str_val(ws, row, C_NOMBRE),
            "cargo":           str_val(ws, row, C_DETALLE),
            "twitter":         str_val(ws, row, C_TW),
            "twitter_activo":  fmt_date(v(ws, row, C_TW_A)),
            "bluesky":         str_val(ws, row, C_BS),
            "bluesky_activo":  fmt_date(v(ws, row, C_BS_A)),
            "mastodon":        str_val(ws, row, C_MD),
            "mastodon_activo": fmt_date(v(ws, row, C_MD_A)),
            "email":           str_val(ws, row, C_EMAIL),
        })
    return result


def build_congreso(ws, rows):
    """Congreso → congreso.json"""
    result = []
    for row in rows:
        result.append({
            "tipo":            str_val(ws, row, C_TIPO),
            "nombre":          str_val(ws, row, C_NOMBRE),
            "grupo":           str_val(ws, row, C_GRUPO),
            "circunscripcion": str_val(ws, row, C_CIRCUN),
            "twitter":         str_val(ws, row, C_TW),
            "twitter_activo":  fmt_date(v(ws, row, C_TW_A)),
            "bluesky":         str_val(ws, row, C_BS),
            "bluesky_activo":  fmt_date(v(ws, row, C_BS_A)),
            "mastodon":        str_val(ws, row, C_MD),
            "mastodon_activo": fmt_date(v(ws, row, C_MD_A)),
            "email":           str_val(ws, row, C_EMAIL),
        })
    return result


def build_senado(ws, rows):
    """Senado → senado.json"""
    result = []
    for row in rows:
        result.append({
            "tipo":            str_val(ws, row, C_TIPO),
            "nombre":          str_val(ws, row, C_NOMBRE),
            "grupo":           str_val(ws, row, C_GRUPO),
            "twitter":         str_val(ws, row, C_TW),
            "twitter_activo":  fmt_date(v(ws, row, C_TW_A)),
            "bluesky":         str_val(ws, row, C_BS),
            "bluesky_activo":  fmt_date(v(ws, row, C_BS_A)),
            "mastodon":        str_val(ws, row, C_MD),
            "mastodon_activo": fmt_date(v(ws, row, C_MD_A)),
            "email":           str_val(ws, row, C_EMAIL),
        })
    return result


def build_partidos(ws, rows):
    """Partidos → partidos.json"""
    result = []
    for row in rows:
        result.append({
            "nombre":          str_val(ws, row, C_NOMBRE),
            "ambito":          "Nacional",
            "twitter":         str_val(ws, row, C_TW),
            "twitter_activo":  fmt_date(v(ws, row, C_TW_A)),
            "bluesky":         str_val(ws, row, C_BS),
            "bluesky_activo":  fmt_date(v(ws, row, C_BS_A)),
            "mastodon":        str_val(ws, row, C_MD),
            "mastodon_activo": fmt_date(v(ws, row, C_MD_A)),
            "email":           str_val(ws, row, C_EMAIL),
        })
    return result


def build_universidades(ws, rows):
    """Universidades → universidades.json
    El tipo (Pública/Privada) no está en el Excel; se lee de C_TIPO si existe,
    o se deja None para que se complete manualmente.
    """
    result = []
    for row in rows:
        tipo = str_val(ws, row, C_TIPO) or str_val(ws, row, C_DETALLE)
        result.append({
            "nombre":          str_val(ws, row, C_NOMBRE),
            "tipo":            tipo,
            "twitter":         str_val(ws, row, C_TW),
            "twitter_activo":  fmt_date(v(ws, row, C_TW_A)),
            "bluesky":         str_val(ws, row, C_BS),
            "bluesky_activo":  fmt_date(v(ws, row, C_BS_A)),
            "mastodon":        str_val(ws, row, C_MD),
            "mastodon_activo": fmt_date(v(ws, row, C_MD_A)),
            "email":           str_val(ws, row, C_EMAIL),
        })
    return result


def build_total(ws, all_rows_by_cat, refs: dict):
    """
    total.json — versión agregada con flags booleanos para _activo.
    Usa las fechas de lastUpdate.json como referencia temporal, no datetime.now().
    subcategoria: col 10 (Detalle) si tiene valor, si no col 11 (Tipo).
    """
    def ref_date(platform: str):
        val = refs.get(platform)
        if not val:
            return None
        try:
            return datetime.strptime(val[:10], "%Y-%m-%d").date()
        except ValueError:
            return None

    tw_ref = ref_date("twitter")
    bs_ref = ref_date("bluesky")
    md_ref = ref_date("mastodon")

    result = []
    for cat, rows in all_rows_by_cat.items():
        for row in rows:
            sub = str_val(ws, row, C_DETALLE) or str_val(ws, row, C_TIPO)
            tw_val   = v(ws, row, C_TW_A)
            bs_val   = v(ws, row, C_BS_A)
            md_val   = v(ws, row, C_MD_A)
            result.append({
                "categoria":       cat,
                "subcategoria":    sub,
                "nombre":          str_val(ws, row, C_NOMBRE),
                "twitter":         str_val(ws, row, C_TW),
                "twitter_activo":  is_active(fmt_date(tw_val), tw_ref),
                "bluesky":         str_val(ws, row, C_BS),
                "bluesky_activo":  is_active(fmt_date(bs_val), bs_ref),
                "mastodon":        str_val(ws, row, C_MD),
                "mastodon_activo": is_active(fmt_date(md_val), md_ref),
            })
    return result


# ── Escritura ──────────────────────────────────────────────────────────────────

def write_json(path: str, data: list, dry_run: bool):
    if dry_run:
        return
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
        f.write("\n")


# ── Main ──────────────────────────────────────────────────────────────────────

CATEGORY_MAP = {
    "AGE":          ("age.json",          build_age),
    "Autonomías":   ("autonomias.json",   build_autonomias),
    "Gobierno":     ("gobierno.json",     build_gobierno),
    "Congreso":     ("congreso.json",     build_congreso),
    "Senado":       ("senado.json",       build_senado),
    "Partidos":     ("partidos.json",     build_partidos),
    "Universidades":("universidades.json",build_universidades),
}


def main():
    dry_run = "--dry-run" in sys.argv

    if dry_run:
        print("MODO DRY-RUN: no se escribirá ningún archivo.\n")

    print(f"Cargando {XLSX}…")
    refs = load_last_update_refs()
    if refs:
        print(f"  Referencias temporales: {refs}")
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb["Sheet1"]

    # Agrupar índices de filas por categoría
    rows_by_cat: dict[str, list[int]] = {cat: [] for cat in CATEGORY_MAP}
    unknown_cats: set = set()

    for row in range(2, ws.max_row + 1):
        cat = ws.cell(row, C_CAT).value
        if cat is None:
            continue
        if cat in CATEGORY_MAP:
            rows_by_cat[cat].append(row)
        else:
            unknown_cats.add(cat)

    if unknown_cats:
        print(f"  ⚠ Categorías desconocidas (se ignoran): {sorted(unknown_cats)}")

    # Generar y escribir cada archivo JSON
    print()
    for cat, (filename, builder) in CATEGORY_MAP.items():
        rows = rows_by_cat[cat]
        data = builder(ws, rows)
        dest = os.path.join(DATA_DIR, filename)
        action = "  (sin cambios)" if dry_run else f"  → {dest}"
        print(f"  {filename}: {len(data)} entradas{action}")
        write_json(dest, data, dry_run)

    # Generar total.json
    total_data = build_total(ws, rows_by_cat, refs)
    total_path = os.path.join(DATA_DIR, "total.json")
    total_action = "  (sin cambios)" if dry_run else f"  → {total_path}"
    print(f"  total.json: {len(total_data)} entradas{total_action}")
    write_json(total_path, total_data, dry_run)

    print()
    if dry_run:
        print("Dry-run completado. Ejecuta sin --dry-run para escribir los archivos.")
    else:
        print("Hecho. Archivos JSON actualizados en src/data/")

    wb.close()


if __name__ == "__main__":
    main()
