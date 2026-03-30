#!/usr/bin/env python3
"""
check_nitter.py — Actualiza "Twitter Activo" via RSS de instancias Nitter (gratis, sin API key).

Uso:
    python3 check_nitter.py                       # comprueba todos los handles pendientes
    python3 check_nitter.py --handle UNIAuniversidad   # prueba un handle concreto
    python3 check_nitter.py --reset               # borra fechas y vuelve a empezar

Notas:
    - Prueba múltiples instancias Nitter en orden hasta que una responda.
    - Si todas fallan el handle queda sin datos (no sobreescribe una fecha ya guardada).
    - Las instancias públicas pueden caerse sin previo aviso; actualiza NITTER_INSTANCES
      con instancias activas si ves muchos fallos.
"""

import json
import os
import re
import sys
import time
import openpyxl
import requests
from datetime import datetime, timezone
from email.utils import parsedate_to_datetime
from xml.etree import ElementTree as ET

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
ROOT_DIR   = os.path.dirname(SCRIPT_DIR)
XLSX       = os.path.join(ROOT_DIR, "datosfinales.xlsx")

COL_NOMBRE = 2
COL_TW     = 3
COL_TW_A   = 4
DELAY      = 0.6   # segundos entre peticiones para no saturar las instancias

# Instancias Nitter públicas — ordenadas por fiabilidad histórica
# Si una deja de funcionar, elimínala o muévela al final
NITTER_INSTANCES = [
    "nitter.privacydev.net",
    "nitter.poast.org",
    "nitter.cz",
    "nitter.net",
    "nitter.1d4.us",
    "nitter.kavin.rocks",
    "nitter.moomoo.me",
]


def fmt(dt: datetime) -> str:
    return dt.strftime("%Y-%m-%d")


def is_done(val) -> bool:
    if isinstance(val, datetime):
        return True
    if isinstance(val, str):
        return bool(re.match(r"^\d{4}-\d{2}-\d{2}", val)) or val == "404"
    return False


def nitter_last_tweet(session: requests.Session, handle: str, verbose: bool = False):
    """
    Intenta obtener la fecha del último tweet via RSS de múltiples instancias Nitter.

    Devuelve:
        'YYYY-MM-DD'  → fecha del último tweet
        '404'         → cuenta no encontrada o suspendida
        None          → todas las instancias fallaron (error transitorio)
    """
    for instance in NITTER_INSTANCES:
        url = f"https://{instance}/{handle}/rss"
        try:
            r = session.get(url, timeout=8)

            if r.status_code == 404:
                if verbose:
                    print(f"    [{instance}] 404")
                return "404"

            if r.status_code != 200:
                if verbose:
                    print(f"    [{instance}] HTTP {r.status_code}, probando siguiente…")
                continue

            # Algunos nitter devuelven HTML de error con 200
            if "<title>Error</title>" in r.text or "Instance is down" in r.text:
                if verbose:
                    print(f"    [{instance}] instancia caída, probando siguiente…")
                continue

            root = ET.fromstring(r.text)
            items = root.findall(".//item")

            if not items:
                # Feed vacío — cuenta existe pero sin tweets
                if verbose:
                    print(f"    [{instance}] feed vacío")
                return None

            pub = items[0].findtext("pubDate")
            if pub:
                dt = parsedate_to_datetime(pub)
                if verbose:
                    print(f"    [{instance}] OK → {fmt(dt)}")
                return fmt(dt)

        except ET.ParseError:
            if verbose:
                print(f"    [{instance}] XML inválido, probando siguiente…")
            continue
        except requests.RequestException as e:
            if verbose:
                print(f"    [{instance}] error de red: {e}, probando siguiente…")
            continue

    return None  # todas las instancias fallaron


def save_last_update():
    lup = os.path.join(ROOT_DIR, "src", "data", "lastUpdate.json")
    try:
        with open(lup, encoding="utf-8") as f:
            data = json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        data = {}
    data["twitter"] = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    with open(lup, "w", encoding="utf-8") as f:
        json.dump(data, f)
    print(f"Fecha de última actualización guardada en {lup}")


def main():
    args = sys.argv[1:]
    handle_only = None
    reset = False

    i = 0
    while i < len(args):
        if args[i] == "--handle" and i + 1 < len(args):
            handle_only = args[i + 1]
            i += 2
        elif args[i] == "--reset":
            reset = True
            i += 1
        else:
            i += 1

    session = requests.Session()
    session.headers.update({
        "User-Agent": "Mozilla/5.0 (compatible; RSS reader)",
        "Accept": "application/rss+xml, application/xml, text/xml",
    })

    # ── Modo --handle: prueba rápida de un handle concreto ────────────────────
    if handle_only:
        print(f"Probando @{handle_only}…")
        result = nitter_last_tweet(session, handle_only, verbose=True)
        if result == "404":
            print(f"\n  @{handle_only}: cuenta no encontrada (404)")
        elif result:
            print(f"\n  @{handle_only}: último tweet el {result}")
        else:
            print(f"\n  @{handle_only}: sin datos (todas las instancias fallaron)")
        return

    # ── Modo normal / reset ───────────────────────────────────────────────────
    print(f"Cargando {XLSX}…")
    wb = openpyxl.load_workbook(XLSX)
    ws = wb["Sheet1"]

    if reset:
        cleared = 0
        for row in range(2, ws.max_row + 1):
            if ws.cell(row, COL_TW).value and is_done(ws.cell(row, COL_TW_A).value):
                ws.cell(row, COL_TW_A).value = None
                cleared += 1
        wb.save(XLSX)
        print(f"Reset: {cleared} fechas borradas. Vuelve a ejecutar sin --reset.")
        return

    done = skipped = errors = no_data = 0

    try:
        for row in range(2, ws.max_row + 1):
            handle = ws.cell(row, COL_TW).value
            nombre = ws.cell(row, COL_NOMBRE).value or f"fila {row}"

            if not handle:
                continue

            if is_done(ws.cell(row, COL_TW_A).value):
                skipped += 1
                continue

            result = nitter_last_tweet(session, handle)
            ws.cell(row, COL_TW_A).value = result

            if result == "404":
                print(f"  @{handle} ({nombre}): no encontrado (404)")
                errors += 1
            elif result:
                print(f"  @{handle} ({nombre}): {result}")
                done += 1
            else:
                print(f"  @{handle} ({nombre}): sin datos")
                no_data += 1

            wb.save(XLSX)
            time.sleep(DELAY)

    finally:
        save_last_update()

    print(f"\nCompletado: {done} con fecha, {errors} no encontrados, "
          f"{no_data} sin datos, {skipped} ya tenían fecha.")
    print("Hecho.")


if __name__ == "__main__":
    main()
