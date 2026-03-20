#!/usr/bin/env python3
"""
check_twitter.py — Actualiza la columna "Twitter Activo" con la fecha del
último tweet usando GetXAPI (https://www.getxapi.com).

Requisitos:
    pip install requests openpyxl

Coste estimado: ~$0.001 por handle (~$0.59 para las 590 cuentas).
GetXAPI ofrece $0.50 en créditos gratuitos al registrarse (sin tarjeta).
Obtén tu clave en: https://www.getxapi.com

Uso:
    python3 check_twitter.py --token TU_API_KEY      # run normal
    python3 check_twitter.py --token TU_API_KEY --discover   # prueba 1 handle y muestra JSON crudo
    python3 check_twitter.py --token TU_API_KEY --reset      # borra fechas ya guardadas (vuelve a empezar)

Resultado en la celda "Twitter Activo":
    "YYYY-MM-DD"  → fecha del último tweet encontrado
    "404"         → cuenta no encontrada o suspendida
    vacío         → sin handle

Re-ejecución:
    Las filas que ya tienen una fecha o "404" se saltan automáticamente.
    Ejecuta cuantas veces necesites hasta completar las 590 cuentas.
"""

import os
import re
import sys
import json
import time
import openpyxl
import requests
from datetime import datetime, timezone

# ── Configuración ─────────────────────────────────────────────────────────────

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
ROOT_DIR   = os.path.dirname(SCRIPT_DIR)
XLSX       = os.path.join(ROOT_DIR, "datosfinales.xlsx")
API_BASE   = "https://api.getxapi.com"

# Columnas (1-indexed)
COL_NOMBRE = 2
COL_TW     = 3
COL_TW_A   = 4

# Campos de fecha en el objeto tweet
DATE_FIELDS = ["createdAt", "created_at", "date", "timestamp", "tweetCreatedAt"]

# Pausa entre llamadas (segundos) — evita rate limiting
DELAY = 0.3


# ── Utilidades ────────────────────────────────────────────────────────────────

def fmt(dt: datetime) -> str:
    return dt.strftime("%Y-%m-%d")


def parse_date(raw):
    """
    Parsea distintos formatos de fecha que puede devolver la API.
    Devuelve 'YYYY-MM-DD' o None.
    """
    if not raw:
        return None
    if isinstance(raw, datetime):
        return fmt(raw)
    s = str(raw).strip()
    # ISO 8601: "2025-02-10T14:30:00.000Z"
    try:
        dt = datetime.fromisoformat(s.replace("Z", "+00:00"))
        return fmt(dt)
    except ValueError:
        pass
    # Twitter legacy: "Mon Feb 10 14:30:00 +0000 2025"
    try:
        from email.utils import parsedate_to_datetime
        dt = parsedate_to_datetime(s)
        return fmt(dt)
    except Exception:
        pass
    # Solo fecha: "2025-02-10"
    if re.match(r"^\d{4}-\d{2}-\d{2}", s):
        return s[:10]
    return None


def extract_date(item):
    """Busca la fecha del tweet en los campos conocidos."""
    for field in DATE_FIELDS:
        val = item.get(field)
        if val:
            parsed = parse_date(val)
            if parsed:
                return parsed
    return None


def is_done(val) -> bool:
    """True si la celda ya tiene un resultado válido (fecha o '404')."""
    if isinstance(val, datetime):
        return True
    if isinstance(val, str):
        return bool(re.match(r"^\d{4}-\d{2}-\d{2}", val)) or val == "404"
    return False


def count_pending(ws) -> int:
    return sum(
        1 for row in range(2, ws.max_row + 1)
        if ws.cell(row, COL_TW).value and not is_done(ws.cell(row, COL_TW_A).value)
    )


# ── GetXAPI ───────────────────────────────────────────────────────────────────

def get_last_tweet_date(session, handle):
    """
    Llama a GetXAPI /twitter/user/tweets y devuelve la fecha del último tweet,
    '404' si la cuenta no existe/está suspendida, o None si hay error transitorio.
    """
    url = f"{API_BASE}/twitter/user/tweets"
    try:
        resp = session.get(url, params={"userName": handle}, timeout=30)

        if resp.status_code == 429:
            # Rate limit — esperar y reintentar una vez
            time.sleep(10)
            resp = session.get(url, params={"userName": handle}, timeout=30)

        data = resp.json()

        # Cuenta no encontrada / suspendida
        if "error" in data:
            err = str(data["error"]).lower()
            if "resolve" in err or "not found" in err or "suspend" in err \
                    or "does not exist" in err or "404" in err:
                return "404"
            # Otro error (transitorio)
            return None

        tweets = data.get("tweets", [])
        if not tweets:
            # Sin tweets pero cuenta existe — devolvemos None (no "404")
            return None

        return extract_date(tweets[0])

    except (requests.RequestException, ValueError):
        return None


# ── Modo discover ─────────────────────────────────────────────────────────────

def discover_mode(session, handle):
    """Llama a la API con un handle y muestra el JSON crudo del primer tweet."""
    print(f"\n── DISCOVER MODE: @{handle} ──")
    url = f"{API_BASE}/twitter/user/tweets"
    try:
        resp = session.get(url, params={"userName": handle}, timeout=30)
        data = resp.json()
        if "error" in data:
            print(f"  Error: {data['error']}")
            return
        tweets = data.get("tweets", [])
        if not tweets:
            print("  (sin tweets — cuenta existe pero vacía)")
            return
        print(f"\nPrimer tweet ({len(tweets)} tweets en la página):")
        print(json.dumps(tweets[0], indent=2, ensure_ascii=False, default=str))
        print(f"\n→ Fecha extraída: {extract_date(tweets[0])}")
    except Exception as e:
        print(f"  Error: {e}")


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    args = sys.argv[1:]

    token = None
    discover = False
    reset = False

    i = 0
    while i < len(args):
        if args[i] == "--token" and i + 1 < len(args):
            token = args[i + 1]
            i += 2
        elif args[i] == "--discover":
            discover = True
            i += 1
        elif args[i] == "--reset":
            reset = True
            i += 1
        else:
            i += 1

    if not token:
        print("Error: falta --token TU_GETXAPI_KEY")
        print("Obtén tu clave en: https://www.getxapi.com")
        sys.exit(1)

    session = requests.Session()
    session.headers.update({"Authorization": f"Bearer {token}"})

    # -- Modo discover --
    if discover:
        discover_mode(session, "sanchezcastejon")
        return

    print(f"Cargando {XLSX}…")
    wb = openpyxl.load_workbook(XLSX)
    ws = wb["Sheet1"]

    # -- Modo reset --
    if reset:
        cleared = 0
        for row in range(2, ws.max_row + 1):
            if ws.cell(row, COL_TW).value and is_done(ws.cell(row, COL_TW_A).value):
                ws.cell(row, COL_TW_A).value = None
                cleared += 1
        wb.save(XLSX)
        print(f"Reset: {cleared} fechas borradas. Vuelve a ejecutar sin --reset.")
        return

    # -- Run normal --
    pending = count_pending(ws)
    print(f"{pending} handles pendientes de comprobar.")
    if pending == 0:
        print("Todo ya está comprobado. Usa --reset para volver a empezar.")
        return

    done = skipped = errors = no_data = 0

    try:
        for row in range(2, ws.max_row + 1):
            handle = ws.cell(row, COL_TW).value
            nombre = ws.cell(row, COL_NOMBRE).value or f"fila {row}"

            if not handle:
                ws.cell(row, COL_TW_A).value = None
                continue

            if is_done(ws.cell(row, COL_TW_A).value):
                skipped += 1
                continue

            result = get_last_tweet_date(session, handle)
            ws.cell(row, COL_TW_A).value = result

            if result == "404":
                print(f"  @{handle} ({nombre}): no encontrado (404)")
                errors += 1
            elif result:
                print(f"  @{handle} ({nombre}): {result}")
                done += 1
            else:
                print(f"  @{handle} ({nombre}): sin datos")
                no_data += 1

            wb.save(XLSX)
            time.sleep(DELAY)
    finally:
        last_update_path = os.path.join(ROOT_DIR, "src", "data", "lastUpdate.json")
        try:
            with open(last_update_path, encoding="utf-8") as f:
                data = json.load(f)
        except (FileNotFoundError, json.JSONDecodeError):
            data = {}
        data["twitter"] = datetime.now(timezone.utc).strftime("%Y-%m-%d")
        with open(last_update_path, "w", encoding="utf-8") as f:
            json.dump(data, f)
        print(f"Fecha de última actualización guardada en {last_update_path}")

    print(f"\nCompletado: {done} con fecha, {errors} no encontrados, "
          f"{no_data} sin datos, {skipped} ya tenían fecha.")
    print("Hecho.")


if __name__ == "__main__":
    main()
